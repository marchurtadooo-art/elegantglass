"""GLASSWORK backend — FastAPI + MongoDB.

Construction site management for aluminum & glass companies.
JWT auth with bcrypt, role-based access (ADMIN, MANAGER, WORKER).
"""
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import logging
import uuid
import bcrypt
import jwt
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal, Any

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr

from security import (
    SecurityHeadersMiddleware,
    ensure_security_indices,
    cleanup_expired_security_records,
    is_login_blocked,
    record_login_attempt,
    reset_login_attempts,
    register_session,
    touch_session,
    is_session_idle,
    blacklist_token,
    is_token_blacklisted,
    audit_log,
    get_client_ip,
    SESSION_IDLE_MINUTES,
    LOGIN_BLOCK_MINUTES,
)
from notifications import (
    DEFAULT_NOTIF_PREFS,
    PREF_KEYS,
    ensure_notification_indices,
    register_push_token,
    unregister_push_token,
    unregister_user_tokens,
    send_push,
)

# ---------- Logging ----------
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("glasswork")

# ---------- Mongo ----------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

# ---------- App / router ----------
app = FastAPI(title="GLASSWORK API")
api = APIRouter(prefix="/api")

JWT_ALGO = "HS256"
JWT_SECRET = os.environ["JWT_SECRET"]
ACCESS_MIN = 15
REFRESH_DAYS = 7

bearer_scheme = HTTPBearer(auto_error=False)


# =========================================================
# Helpers
# =========================================================
def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def iso(dt: Optional[datetime]) -> Optional[str]:
    if not dt:
        return None
    if isinstance(dt, str):
        return dt
    return dt.replace(tzinfo=timezone.utc).isoformat() if dt.tzinfo is None else dt.isoformat()


def hash_password(plain: str) -> str:
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_token(payload: dict, expires_delta: timedelta) -> str:
    to_encode = payload.copy()
    to_encode["exp"] = now_utc() + expires_delta
    to_encode["iat"] = now_utc()
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGO)


def create_access_token(user_id: str, role: str, company_id: str, jti: Optional[str] = None) -> str:
    payload = {"sub": user_id, "role": role, "cid": company_id, "type": "access"}
    if jti:
        payload["jti"] = jti
    return create_token(payload, timedelta(minutes=ACCESS_MIN))


def create_refresh_token(user_id: str) -> str:
    return create_token({"sub": user_id, "type": "refresh"}, timedelta(days=REFRESH_DAYS))


def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


def serialize(doc: dict) -> dict:
    """Strip _id and convert datetimes to ISO."""
    if not doc:
        return doc
    doc.pop("_id", None)
    for k, v in list(doc.items()):
        if isinstance(v, datetime):
            doc[k] = iso(v)
    return doc


async def get_current_user(
    request: Request,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(bearer_scheme),
) -> dict:
    if credentials is None or not credentials.credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = credentials.credentials
    payload = decode_token(token)
    if payload.get("type") != "access":
        raise HTTPException(status_code=401, detail="Invalid token type")

    # Security layer: token blacklist (logged out tokens)
    if await is_token_blacklisted(db, token):
        raise HTTPException(status_code=401, detail="Token revocado, inicia sesión de nuevo")

    # Security layer: session inactivity (>30min idle)
    jti = payload.get("jti")
    if jti:
        session = await touch_session(db, jti)
        if session is not None:
            idle, _mins = await is_session_idle(session)
            if idle:
                # Idle timeout — blacklist this token & require new login
                await db.token_sessions.delete_one({"jti": jti})
                await blacklist_token(db, token, jti, None)
                raise HTTPException(
                    status_code=401,
                    detail=f"Sesión expirada por inactividad ({SESSION_IDLE_MINUTES} min). Inicia sesión de nuevo.",
                )

    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    # Stash the raw token + jti so endpoints can blacklist on logout / audit
    request.state.access_token = token
    request.state.access_jti = jti
    return user


def require_role(*roles: str):
    async def _dep(user: dict = Depends(get_current_user)) -> dict:
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return _dep


def strip_costs_for_worker(obj: dict, role: str) -> dict:
    """Hide ALL financial fields from any role that is NOT ADMIN.
    Function name kept for backwards-compat — the rule changed to ADMIN-only
    visibility for budget, spent, remaining, unit_price, total_cost,
    projected_final_cost and value."""
    if role == "ADMIN":
        return obj
    out = dict(obj)
    for k in (
        "budget", "spent", "remaining",
        "total_cost", "unit_price", "projected_final_cost", "value",
    ):
        if k in out:
            out[k] = None
    return out


# =========================================================
# Models
# =========================================================
Role = Literal["ADMIN", "MANAGER", "WORKER"]
ProjectStatus = Literal["PENDING", "ACTIVE", "PAUSED", "COMPLETED", "CANCELLED"]
MatCategory = Literal["PERFILERIA", "VIDRIO", "HERRAJES", "SELLANTES", "HERRAMIENTAS", "CONSUMIBLES", "OTROS"]
EntryType = Literal["PURCHASE", "USAGE", "RETURN", "ADJUSTMENT"]
PhotoType = Literal["PROGRESS", "BEFORE", "AFTER", "INCIDENT", "MATERIAL", "MEASUREMENT"]
Weather = Literal["SUNNY", "CLOUDY", "RAINY", "WINDY", "STOPPED_BY_WEATHER"]
LogStatus = Literal["PENDING", "APPROVED", "REJECTED"]
AlertType = Literal["LOW_STOCK", "BUDGET_EXCEEDED", "PROJECT_DELAYED", "INCIDENT_REPORTED", "LOG_MISSING"]
AlertSeverity = Literal["INFO", "WARNING", "CRITICAL"]


class RegisterIn(BaseModel):
    name: str
    email: EmailStr
    password: str
    company_name: str
    phone: Optional[str] = None


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class RefreshIn(BaseModel):
    refresh_token: str


class TokenOut(BaseModel):
    access_token: str
    refresh_token: str
    user: dict


class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: Role = "WORKER"
    phone: Optional[str] = None


class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[Role] = None
    phone: Optional[str] = None
    is_active: Optional[bool] = None


class ProfileUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None


class CompanyUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    logo: Optional[str] = None  # base64


class ForgotPasswordIn(BaseModel):
    email: EmailStr


class ProjectIn(BaseModel):
    name: str
    description: Optional[str] = ""
    address: str
    lat: Optional[float] = None
    lng: Optional[float] = None
    status: ProjectStatus = "PENDING"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    actual_end_date: Optional[str] = None
    budget: float = 0.0
    client_name: str = ""
    client_phone: str = ""
    client_email: str = ""
    notes: str = ""
    assigned_worker_ids: List[str] = []
    cover_photo: Optional[str] = None


class MaterialIn(BaseModel):
    name: str
    unit: Literal["m2", "m", "ud", "kg", "l", "caja"]
    category: MatCategory
    unit_price: float = 0.0
    supplier: str = ""
    notes: str = ""


class MaterialEntryIn(BaseModel):
    project_id: str
    material_id: str
    quantity: float
    unit_price: Optional[float] = None  # if None, use catalog price
    type: EntryType = "USAGE"
    date: Optional[str] = None
    notes: str = ""
    receipt_photo: Optional[str] = None  # base64


class DailyLogIn(BaseModel):
    project_id: str
    date: Optional[str] = None
    hours_worked: float
    work_description: str
    weather_condition: Weather = "SUNNY"
    progress_percentage: float = 0.0
    incidents: Optional[str] = None
    photo_ids: List[str] = []
    material_entry_ids: List[str] = []


class LogReviewIn(BaseModel):
    status: Literal["APPROVED", "REJECTED"]
    review_comment: Optional[str] = ""


class PhotoIn(BaseModel):
    project_id: str
    daily_log_id: Optional[str] = None
    image_base64: str  # full data URI or just base64
    caption: str = ""
    photo_type: PhotoType = "PROGRESS"


class PushTokenIn(BaseModel):
    token: str
    platform: Optional[str] = "unknown"  # 'ios' | 'android' | 'web'


class PushTokenDeleteIn(BaseModel):
    token: str


class NotificationPreferencesIn(BaseModel):
    new_alert: Optional[bool] = None
    new_project: Optional[bool] = None
    log_approved: Optional[bool] = None
    log_rejected: Optional[bool] = None
    incident_reported: Optional[bool] = None
    budget_exceeded: Optional[bool] = None


class AlertCreate(BaseModel):
    type: AlertType
    severity: AlertSeverity = "INFO"
    message: str
    project_id: str  # mandatory


# =========================================================
# Auth endpoints
# =========================================================
@api.get("/")
async def api_root():
    """Lightweight root — used by frontend for warm-up / health checks."""
    return {"app": "GLASSWORK", "ok": True, "service": "glasswork-api", "ts": iso(now_utc())}


@api.get("/health")
async def api_health():
    """Health check used by frontend for backend warm-up before login.
    Verifies DB connectivity quickly (best-effort)."""
    db_ok = True
    try:
        await db.command("ping")
    except Exception:
        db_ok = False
    return {"ok": True, "db": db_ok, "ts": iso(now_utc())}


@api.post("/auth/register", response_model=TokenOut)
async def register(body: RegisterIn):
    email = body.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email ya registrado")
    company_id = str(uuid.uuid4())
    company = {
        "id": company_id,
        "name": body.company_name,
        "logo": None,
        "address": "",
        "phone": "",
        "email": "",
        "created_at": now_utc(),
    }
    await db.companies.insert_one(company.copy())
    user_id = str(uuid.uuid4())
    user = {
        "id": user_id,
        "name": body.name,
        "email": email,
        "password_hash": hash_password(body.password),
        "role": "ADMIN",
        "company_id": company_id,
        "phone": body.phone or "",
        "avatar": None,
        "is_active": True,
        "notification_preferences": dict(DEFAULT_NOTIF_PREFS),
        "created_at": now_utc(),
        "updated_at": now_utc(),
    }
    await db.users.insert_one(user.copy())
    access = create_access_token(user_id, "ADMIN", company_id)
    refresh = create_refresh_token(user_id)
    user_out = {k: v for k, v in user.items() if k != "password_hash"}
    return TokenOut(access_token=access, refresh_token=refresh, user=serialize(user_out))


@api.post("/auth/login", response_model=TokenOut)
async def login(body: LoginIn, request: Request):
    email = body.email.lower().strip()
    ip = get_client_ip(request)

    # 1) Lockout check — 5 fails in 15 min
    blocked, remaining = await is_login_blocked(db, email)
    if blocked:
        await audit_log(
            db, action="LOGIN_BLOCKED", resource="auth", request=request,
            user_email=email, success=False,
            extra={"remaining_minutes": remaining},
        )
        raise HTTPException(
            status_code=429,
            detail=f"Demasiados intentos, espera {remaining} minutos",
        )

    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        # Record failed attempt + audit
        await record_login_attempt(db, email, ip, success=False)
        await audit_log(
            db, action="LOGIN_FAILED", resource="auth", request=request,
            user_email=email, success=False,
        )
        # Re-check lockout to inform user precisely
        blocked2, remaining2 = await is_login_blocked(db, email)
        if blocked2:
            raise HTTPException(
                status_code=429,
                detail=f"Demasiados intentos, espera {remaining2} minutos",
            )
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    if not user.get("is_active", True):
        await audit_log(
            db, action="LOGIN_FAILED", resource="auth", request=request,
            user=user, success=False, extra={"reason": "account_disabled"},
        )
        raise HTTPException(status_code=403, detail="Cuenta desactivada")

    # Successful login → issue tokens with jti, register session, reset failures
    jti = str(uuid.uuid4())
    access = create_access_token(user["id"], user["role"], user["company_id"], jti=jti)
    refresh = create_refresh_token(user["id"])
    exp_dt = now_utc() + timedelta(minutes=ACCESS_MIN)
    await register_session(db, jti, user["id"], exp_dt, ip)
    await record_login_attempt(db, email, ip, success=True)
    await reset_login_attempts(db, email)
    await audit_log(
        db, action="LOGIN", resource="auth", request=request,
        user=user, success=True, extra={"jti": jti},
    )

    user_out = {k: v for k, v in user.items() if k not in ("_id", "password_hash")}
    return TokenOut(access_token=access, refresh_token=refresh, user=serialize(user_out))


@api.post("/auth/refresh")
async def refresh_token(body: RefreshIn, request: Request):
    payload = decode_token(body.refresh_token)
    if payload.get("type") != "refresh":
        raise HTTPException(status_code=401, detail="Invalid token type")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    jti = str(uuid.uuid4())
    access = create_access_token(user["id"], user["role"], user["company_id"], jti=jti)
    exp_dt = now_utc() + timedelta(minutes=ACCESS_MIN)
    await register_session(db, jti, user["id"], exp_dt, get_client_ip(request))
    return {"access_token": access}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return serialize(user)


@api.post("/auth/logout")
async def logout(request: Request, user: dict = Depends(get_current_user)):
    # Blacklist the access token used for this request
    token = getattr(request.state, "access_token", None)
    jti = getattr(request.state, "access_jti", None)
    if token:
        # We use ACCESS_MIN as expiry estimate; the entry will be cleaned up on startup
        await blacklist_token(db, token, jti, now_utc() + timedelta(minutes=ACCESS_MIN))
    # Remove this device's push token (best-effort — caller may pass it via body)
    try:
        body = await request.json()
        if isinstance(body, dict) and body.get("push_token"):
            await unregister_push_token(db, body["push_token"])
    except Exception:
        pass
    await audit_log(
        db, action="LOGOUT", resource="auth", request=request,
        user=user, success=True, extra={"jti": jti},
    )
    return {"ok": True}


# =========================================================
# Users / team management
# =========================================================
@api.get("/users")
async def list_users(user: dict = Depends(get_current_user)):
    users = await db.users.find({"company_id": user["company_id"]}, {"_id": 0, "password_hash": 0}).to_list(500)
    # add stats: total hours this month, logs this month, projects assigned
    month_start = datetime(now_utc().year, now_utc().month, 1, tzinfo=timezone.utc)
    out = []
    for u in users:
        logs = await db.daily_logs.find({"worker_id": u["id"], "date": {"$gte": month_start}}).to_list(2000)
        u["hours_this_month"] = round(sum(l.get("hours_worked", 0) for l in logs), 1)
        u["logs_this_month"] = len(logs)
        proj_count = await db.projects.count_documents({"assigned_worker_ids": u["id"]})
        u["projects_count"] = proj_count
        out.append(serialize(u))
    return out


@api.post("/users")
async def create_user(body: UserCreate, user: dict = Depends(require_role("ADMIN"))):
    email = body.email.lower().strip()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email ya registrado")
    nu = {
        "id": str(uuid.uuid4()),
        "name": body.name,
        "email": email,
        "password_hash": hash_password(body.password),
        "role": body.role,
        "company_id": user["company_id"],
        "phone": body.phone or "",
        "avatar": None,
        "is_active": True,
        "notification_preferences": dict(DEFAULT_NOTIF_PREFS),
        "created_at": now_utc(),
        "updated_at": now_utc(),
    }
    await db.users.insert_one(nu.copy())
    nu.pop("password_hash", None)
    return serialize(nu)


@api.patch("/users/{user_id}")
async def update_user(request: Request, user_id: str, body: UserUpdate, user: dict = Depends(require_role("ADMIN"))):
    prev = await db.users.find_one({"id": user_id, "company_id": user["company_id"]}, {"_id": 0, "password_hash": 0})
    upd = {k: v for k, v in body.dict().items() if v is not None}
    upd["updated_at"] = now_utc()
    await db.users.update_one({"id": user_id, "company_id": user["company_id"]}, {"$set": upd})
    u = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    # Audit role change specifically
    if prev and "role" in upd and (prev.get("role") != upd.get("role")):
        await audit_log(
            db, action="USER_ROLE_CHANGE", resource="user", resource_id=user_id,
            request=request, user=user,
            extra={
                "target_email": prev.get("email"),
                "previous_role": prev.get("role"),
                "new_role": upd.get("role"),
            },
        )
    return serialize(u)


# =========================================================
# Projects
# =========================================================
async def project_metrics(project_id: str) -> dict:
    """Compute spent, photo count, log count."""
    pipeline = [{"$match": {"project_id": project_id}}, {"$group": {"_id": None, "spent": {"$sum": "$total_cost"}}}]
    res = await db.material_entries.aggregate(pipeline).to_list(1)
    spent = res[0]["spent"] if res else 0.0
    photo_count = await db.project_photos.count_documents({"project_id": project_id})
    log_count = await db.daily_logs.count_documents({"project_id": project_id})
    return {"spent": round(spent, 2), "photo_count": photo_count, "log_count": log_count}


def filter_project_for_role(p: dict, role: str) -> dict:
    p = serialize(p)
    if role != "ADMIN":
        for k in ("budget", "spent", "remaining", "projected_final_cost"):
            if k in p:
                p[k] = None
    return p


@api.get("/projects")
async def list_projects(
    status_f: Optional[str] = Query(None, alias="status"),
    user: dict = Depends(get_current_user),
):
    q: dict = {"company_id": user["company_id"]}
    if status_f:
        q["status"] = status_f
    if user["role"] == "WORKER":
        q["assigned_worker_ids"] = user["id"]
    projects = await db.projects.find(q).sort("created_at", -1).to_list(500)
    out = []
    for p in projects:
        m = await project_metrics(p["id"])
        p["spent"] = m["spent"]
        p["remaining"] = round((p.get("budget", 0) - m["spent"]), 2)
        p["photo_count"] = m["photo_count"]
        p["log_count"] = m["log_count"]
        out.append(filter_project_for_role(p, user["role"]))
    return out


@api.get("/projects/{project_id}")
async def get_project(project_id: str, user: dict = Depends(get_current_user)):
    p = await db.projects.find_one({"id": project_id, "company_id": user["company_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if user["role"] == "WORKER" and user["id"] not in p.get("assigned_worker_ids", []):
        raise HTTPException(status_code=403, detail="Sin acceso a este proyecto")
    m = await project_metrics(project_id)
    p["spent"] = m["spent"]
    p["remaining"] = round((p.get("budget", 0) - m["spent"]), 2)
    p["photo_count"] = m["photo_count"]
    p["log_count"] = m["log_count"]
    # assigned workers info
    workers = await db.users.find(
        {"id": {"$in": p.get("assigned_worker_ids", [])}}, {"_id": 0, "password_hash": 0}
    ).to_list(50)
    p["assigned_workers"] = [serialize(w) for w in workers]
    return filter_project_for_role(p, user["role"])


@api.post("/projects")
async def create_project(request: Request, body: ProjectIn, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    pid = str(uuid.uuid4())
    proj = {
        "id": pid,
        "company_id": user["company_id"],
        "manager_id": user["id"],
        "created_at": now_utc(),
        "updated_at": now_utc(),
        **body.dict(),
    }
    await db.projects.insert_one(proj.copy())
    await audit_log(
        db, action="PROJECT_CREATE", resource="project", resource_id=pid,
        request=request, user=user, extra={"name": proj.get("name")},
    )
    # Push notification: new project to all users in company
    try:
        await send_push(
            db,
            company_id=user["company_id"],
            preference_key="new_project",
            title="Nueva obra creada",
            body=f"{proj.get('name', 'Nueva obra')} · {proj.get('address', '')}".strip(" ·"),
            data={"type": "new_project", "project_id": pid},
            exclude_user_id=user["id"],
        )
    except Exception as e:
        logger.warning(f"push new_project failed: {e}")
    return serialize(proj)


@api.patch("/projects/{project_id}")
async def update_project(
    project_id: str, body: ProjectIn, user: dict = Depends(require_role("ADMIN", "MANAGER"))
):
    upd = body.dict()
    upd["updated_at"] = now_utc()
    await db.projects.update_one(
        {"id": project_id, "company_id": user["company_id"]}, {"$set": upd}
    )
    p = await db.projects.find_one({"id": project_id})
    return serialize(p)


@api.delete("/projects/{project_id}")
async def delete_project(request: Request, project_id: str, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    proj = await db.projects.find_one({"id": project_id, "company_id": user["company_id"]}, {"_id": 0, "name": 1})
    await db.projects.delete_one({"id": project_id, "company_id": user["company_id"]})
    await audit_log(
        db, action="PROJECT_DELETE", resource="project", resource_id=project_id,
        request=request, user=user, extra={"name": (proj or {}).get("name")},
    )
    return {"ok": True}


# =========================================================
# Materials catalog
# =========================================================
@api.get("/materials")
async def list_materials(
    category: Optional[str] = None,
    q: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    flt: dict = {"company_id": user["company_id"]}
    if category:
        flt["category"] = category
    if q:
        flt["name"] = {"$regex": q, "$options": "i"}
    mats = await db.materials.find(flt).sort("name", 1).to_list(2000)
    out = []
    for m in mats:
        m = serialize(m)
        if user["role"] != "ADMIN":
            m["unit_price"] = None
        out.append(m)
    return out


@api.post("/materials")
async def create_material(body: MaterialIn, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    m = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "is_archived": False,
        "created_at": now_utc(),
        **body.dict(),
    }
    await db.materials.insert_one(m.copy())
    return serialize(m)


@api.patch("/materials/{mid}")
async def update_material(mid: str, body: MaterialIn, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    upd = body.dict()
    await db.materials.update_one({"id": mid, "company_id": user["company_id"]}, {"$set": upd})
    m = await db.materials.find_one({"id": mid})
    return serialize(m)


# =========================================================
# Material entries
# =========================================================
@api.get("/material-entries")
async def list_material_entries(
    project_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    flt: dict = {"company_id": user["company_id"]}
    if project_id:
        flt["project_id"] = project_id
    if user["role"] == "WORKER":
        flt["worker_id"] = user["id"]
    entries = await db.material_entries.find(flt).sort("date", -1).to_list(1000)
    # enrich with material info
    out = []
    for e in entries:
        e = serialize(e)
        mat = await db.materials.find_one({"id": e["material_id"]}, {"_id": 0})
        e["material"] = serialize(mat) if mat else None
        if user["role"] != "ADMIN":
            e["unit_price"] = None
            e["total_cost"] = None
        out.append(e)
    return out


@api.post("/material-entries")
async def create_material_entry(body: MaterialEntryIn, user: dict = Depends(get_current_user)):
    mat = await db.materials.find_one({"id": body.material_id, "company_id": user["company_id"]})
    if not mat:
        raise HTTPException(status_code=404, detail="Material no encontrado")
    proj = await db.projects.find_one({"id": body.project_id, "company_id": user["company_id"]})
    if not proj:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if user["role"] == "WORKER" and user["id"] not in proj.get("assigned_worker_ids", []):
        raise HTTPException(status_code=403, detail="Sin acceso al proyecto")
    unit_price = body.unit_price if body.unit_price is not None else mat.get("unit_price", 0)
    if user["role"] == "WORKER":
        # workers cannot define prices
        unit_price = mat.get("unit_price", 0)
    total = round(unit_price * body.quantity, 2)
    entry_date = body.date or now_utc().isoformat()
    entry = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "project_id": body.project_id,
        "material_id": body.material_id,
        "quantity": body.quantity,
        "unit_price": unit_price,
        "total_cost": total,
        "type": body.type,
        "worker_id": user["id"],
        "date": entry_date,
        "notes": body.notes,
        "receipt_photo": body.receipt_photo,
        "created_at": now_utc(),
    }
    await db.material_entries.insert_one(entry.copy())
    out = serialize(entry)
    if user["role"] != "ADMIN":
        out["unit_price"] = None
        out["total_cost"] = None
    return out


# =========================================================
# Daily logs
# =========================================================
@api.get("/daily-logs")
async def list_logs(
    project_id: Optional[str] = None,
    worker_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    flt: dict = {"company_id": user["company_id"]}
    if project_id:
        flt["project_id"] = project_id
    if worker_id:
        flt["worker_id"] = worker_id
    if user["role"] == "WORKER":
        flt["worker_id"] = user["id"]
    logs = await db.daily_logs.find(flt).sort("date", -1).to_list(1000)
    out = []
    for log in logs:
        log = serialize(log)
        worker = await db.users.find_one({"id": log["worker_id"]}, {"_id": 0, "password_hash": 0})
        log["worker"] = serialize(worker) if worker else None
        proj = await db.projects.find_one({"id": log["project_id"]}, {"_id": 0})
        log["project_name"] = (proj or {}).get("name")
        out.append(log)
    return out


@api.get("/daily-logs/{log_id}")
async def get_log(log_id: str, user: dict = Depends(get_current_user)):
    log = await db.daily_logs.find_one({"id": log_id, "company_id": user["company_id"]})
    if not log:
        raise HTTPException(status_code=404, detail="Parte no encontrado")
    if user["role"] == "WORKER" and log["worker_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso")
    log = serialize(log)
    worker = await db.users.find_one({"id": log["worker_id"]}, {"_id": 0, "password_hash": 0})
    log["worker"] = serialize(worker) if worker else None
    photos = await db.project_photos.find({"id": {"$in": log.get("photo_ids", [])}}, {"_id": 0}).to_list(50)
    log["photos"] = [serialize(p) for p in photos]
    return log


@api.post("/daily-logs")
async def create_log(body: DailyLogIn, user: dict = Depends(get_current_user)):
    proj = await db.projects.find_one({"id": body.project_id, "company_id": user["company_id"]})
    if not proj:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if user["role"] == "WORKER" and user["id"] not in proj.get("assigned_worker_ids", []):
        raise HTTPException(status_code=403, detail="Sin acceso al proyecto")
    if len(body.work_description.strip()) < 20:
        raise HTTPException(status_code=400, detail="Descripción mínima 20 caracteres")
    log = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "project_id": body.project_id,
        "worker_id": user["id"],
        "date": body.date or now_utc().isoformat(),
        "hours_worked": body.hours_worked,
        "work_description": body.work_description,
        "weather_condition": body.weather_condition,
        "progress_percentage": body.progress_percentage,
        "incidents": body.incidents,
        "photo_ids": body.photo_ids,
        "material_entry_ids": body.material_entry_ids,
        "status": "PENDING",
        "review_comment": "",
        "submitted_at": now_utc(),
        "approved_by": None,
        "approved_at": None,
    }
    await db.daily_logs.insert_one(log.copy())
    # auto alert if incident
    if body.incidents and body.incidents.strip():
        alert_doc = {
            "id": str(uuid.uuid4()),
            "company_id": user["company_id"],
            "type": "INCIDENT_REPORTED",
            "severity": "WARNING",
            "message": f"Incidente reportado en {proj['name']} por {user['name']}",
            "project_id": body.project_id,
            "is_read": False,
            "created_at": now_utc(),
        }
        await db.alerts.insert_one(alert_doc.copy())
        # Push notification: incident reported → admins/managers in the company
        try:
            mgrs = await db.users.find(
                {"company_id": user["company_id"], "role": {"$in": ["ADMIN", "MANAGER"]}, "is_active": True},
                {"_id": 0, "id": 1},
            ).to_list(500)
            mgr_ids = [m["id"] for m in mgrs]
            await send_push(
                db,
                company_id=user["company_id"],
                preference_key="incident_reported",
                title="Incidente reportado",
                body=alert_doc["message"],
                data={"type": "incident_reported", "project_id": body.project_id, "alert_id": alert_doc["id"]},
                user_ids=mgr_ids,
                exclude_user_id=user["id"],
            )
        except Exception as e:
            logger.warning(f"push incident_reported failed: {e}")
    # update project progress
    if body.progress_percentage and body.progress_percentage > proj.get("progress_percentage", 0):
        await db.projects.update_one(
            {"id": body.project_id},
            {"$set": {"progress_percentage": body.progress_percentage, "updated_at": now_utc()}},
        )
    return serialize(log)


@api.patch("/daily-logs/{log_id}/review")
async def review_log(log_id: str, body: LogReviewIn, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    res = await db.daily_logs.update_one(
        {"id": log_id, "company_id": user["company_id"]},
        {"$set": {
            "status": body.status,
            "review_comment": body.review_comment or "",
            "approved_by": user["id"],
            "approved_at": now_utc(),
        }},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Parte no encontrado")
    log = await db.daily_logs.find_one({"id": log_id}, {"_id": 0})
    # Push notification → worker that submitted the log
    try:
        worker_id = (log or {}).get("worker_id")
        if worker_id:
            proj = await db.projects.find_one({"id": log.get("project_id")}, {"_id": 0, "name": 1})
            pname = (proj or {}).get("name", "obra")
            if body.status == "APPROVED":
                await send_push(
                    db,
                    company_id=user["company_id"],
                    preference_key="log_approved",
                    title="Parte aprobado ✔",
                    body=f"Tu parte de {pname} ha sido aprobado.",
                    data={"type": "log_approved", "log_id": log_id, "project_id": log.get("project_id")},
                    user_ids=[worker_id],
                )
            else:
                await send_push(
                    db,
                    company_id=user["company_id"],
                    preference_key="log_rejected",
                    title="Parte rechazado",
                    body=f"Tu parte de {pname} requiere correcciones." + (f" {body.review_comment}" if body.review_comment else ""),
                    data={"type": "log_rejected", "log_id": log_id, "project_id": log.get("project_id")},
                    user_ids=[worker_id],
                )
    except Exception as e:
        logger.warning(f"push log review failed: {e}")
    return serialize(log)


# =========================================================
# Photos
# =========================================================
@api.get("/photos")
async def list_photos(
    project_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    flt: dict = {"company_id": user["company_id"]}
    if project_id:
        flt["project_id"] = project_id
    if user["role"] == "WORKER":
        flt["worker_id"] = user["id"]
    photos = await db.project_photos.find(flt).sort("taken_at", -1).to_list(500)
    out = []
    for p in photos:
        p = serialize(p)
        worker = await db.users.find_one({"id": p["worker_id"]}, {"_id": 0, "password_hash": 0})
        p["worker"] = serialize(worker) if worker else None
        out.append(p)
    return out


@api.post("/photos")
async def upload_photo(body: PhotoIn, user: dict = Depends(get_current_user)):
    proj = await db.projects.find_one({"id": body.project_id, "company_id": user["company_id"]})
    if not proj:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if user["role"] == "WORKER" and user["id"] not in proj.get("assigned_worker_ids", []):
        raise HTTPException(status_code=403, detail="Sin acceso al proyecto")
    p = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "project_id": body.project_id,
        "daily_log_id": body.daily_log_id,
        "worker_id": user["id"],
        "image_base64": body.image_base64,
        "caption": body.caption,
        "photo_type": body.photo_type,
        "taken_at": now_utc(),
        "uploaded_at": now_utc(),
    }
    await db.project_photos.insert_one(p.copy())
    return serialize(p)


@api.delete("/photos/{pid}")
async def delete_photo(pid: str, user: dict = Depends(get_current_user)):
    photo = await db.project_photos.find_one({"id": pid, "company_id": user["company_id"]})
    if not photo:
        raise HTTPException(status_code=404, detail="Foto no encontrada")
    if user["role"] == "WORKER" and photo["worker_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Sin permiso")
    await db.project_photos.delete_one({"id": pid})
    return {"ok": True}


# =========================================================
# Dashboard / alerts / reports
# =========================================================
@api.get("/dashboard/summary")
async def dashboard_summary(user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    cid = user["company_id"]
    today = datetime(now_utc().year, now_utc().month, now_utc().day, tzinfo=timezone.utc)
    week_start = today - timedelta(days=today.weekday())
    month_start = datetime(now_utc().year, now_utc().month, 1, tzinfo=timezone.utc)
    active = await db.projects.count_documents({"company_id": cid, "status": "ACTIVE"})
    workers_today = await db.daily_logs.distinct(
        "worker_id", {"company_id": cid, "date": {"$gte": today.isoformat()}}
    )
    pending_logs = await db.daily_logs.count_documents({"company_id": cid, "status": "PENDING"})
    open_alerts = await db.alerts.count_documents({"company_id": cid, "is_read": False})
    week_pipeline = [
        {"$match": {"company_id": cid, "created_at": {"$gte": week_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_cost"}}},
    ]
    month_pipeline = [
        {"$match": {"company_id": cid, "created_at": {"$gte": month_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_cost"}}},
    ]
    week_res = await db.material_entries.aggregate(week_pipeline).to_list(1)
    month_res = await db.material_entries.aggregate(month_pipeline).to_list(1)
    week_spend = round(week_res[0]["total"], 2) if week_res else 0.0
    month_spend = round(month_res[0]["total"], 2) if month_res else 0.0
    # spend by project (last 30 days)
    last30 = now_utc() - timedelta(days=30)
    spend_pipe = [
        {"$match": {"company_id": cid, "created_at": {"$gte": last30}}},
        {"$group": {"_id": "$project_id", "total": {"$sum": "$total_cost"}}},
        {"$sort": {"total": -1}},
        {"$limit": 6},
    ]
    spend_by_proj = await db.material_entries.aggregate(spend_pipe).to_list(6)
    chart = []
    for s in spend_by_proj:
        proj = await db.projects.find_one({"id": s["_id"]}, {"_id": 0, "name": 1})
        chart.append({"project": (proj or {}).get("name", "—"), "amount": round(s["total"], 2)})
    # today's photo activity
    today_photos = await db.project_photos.find(
        {"company_id": cid, "uploaded_at": {"$gte": today}}
    ).sort("uploaded_at", -1).to_list(10)
    photo_feed = []
    for ph in today_photos:
        ph = serialize(ph)
        w = await db.users.find_one({"id": ph["worker_id"]}, {"_id": 0, "password_hash": 0})
        proj = await db.projects.find_one({"id": ph["project_id"]}, {"_id": 0, "name": 1})
        photo_feed.append({
            "id": ph["id"],
            "worker_name": (w or {}).get("name"),
            "project_name": (proj or {}).get("name"),
            "photo_type": ph.get("photo_type"),
            "image_base64": ph.get("image_base64"),
            "uploaded_at": ph.get("uploaded_at"),
        })
    return {
        "active_projects": active,
        "workers_today": len(workers_today),
        "week_spend": week_spend,
        "month_spend": month_spend,
        "pending_logs": pending_logs,
        "open_alerts": open_alerts,
        "spend_by_project": chart,
        "photo_feed": photo_feed,
    }


@api.get("/alerts")
async def list_alerts(user: dict = Depends(get_current_user)):
    alerts = await db.alerts.find({"company_id": user["company_id"]}).sort("created_at", -1).to_list(200)
    return [serialize(a) for a in alerts]


@api.post("/alerts")
async def create_alert(
    request: Request,
    body: AlertCreate,
    user: dict = Depends(require_role("ADMIN", "MANAGER")),
):
    msg = (body.message or "").strip()
    if len(msg) < 3:
        raise HTTPException(status_code=400, detail="El mensaje es obligatorio (mín. 3 caracteres)")
    proj = await db.projects.find_one(
        {"id": body.project_id, "company_id": user["company_id"]}, {"_id": 0, "name": 1}
    )
    if not proj:
        raise HTTPException(status_code=404, detail="Obra no encontrada")
    alert_doc = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "type": body.type,
        "severity": body.severity,
        "message": msg,
        "project_id": body.project_id,
        "is_read": False,
        "created_by": user["id"],
        "created_at": now_utc(),
    }
    await db.alerts.insert_one(alert_doc.copy())
    await audit_log(
        db, action="ALERT_CREATE", resource="alert", resource_id=alert_doc["id"],
        request=request, user=user, extra={"type": body.type, "severity": body.severity, "project_id": body.project_id},
    )
    # Push notification to all users in company with new_alert preference
    try:
        sev_emoji = {"CRITICAL": "🚨", "WARNING": "⚠️", "INFO": "ℹ️"}.get(body.severity, "")
        title = f"{sev_emoji} Nueva alerta · {proj.get('name', '')}".strip()
        await send_push(
            db,
            company_id=user["company_id"],
            preference_key="new_alert",
            title=title or "Nueva alerta",
            body=msg,
            data={"type": "new_alert", "alert_id": alert_doc["id"], "project_id": body.project_id, "severity": body.severity},
            exclude_user_id=user["id"],
        )
    except Exception as e:
        logger.warning(f"push new_alert failed: {e}")
    return serialize(alert_doc)


@api.patch("/alerts/{aid}/read")
async def mark_alert_read(aid: str, user: dict = Depends(get_current_user)):
    await db.alerts.update_one({"id": aid, "company_id": user["company_id"]}, {"$set": {"is_read": True}})
    return {"ok": True}


@api.get("/reports/weekly")
async def list_weekly_reports(user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    reports = await db.weekly_reports.find({"company_id": user["company_id"]}).sort("week_start", -1).to_list(50)
    return [serialize(r) for r in reports]


@api.post("/reports/weekly/generate")
async def generate_weekly_report(user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    cid = user["company_id"]
    today = now_utc()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    pipe_spend = [
        {"$match": {"company_id": cid, "created_at": {"$gte": week_start, "$lte": week_end}}},
        {"$group": {"_id": "$project_id", "spend": {"$sum": "$total_cost"}}},
    ]
    spend = await db.material_entries.aggregate(pipe_spend).to_list(100)
    total_spend = round(sum(s["spend"] for s in spend), 2)
    log_count = await db.daily_logs.count_documents(
        {"company_id": cid, "submitted_at": {"$gte": week_start, "$lte": week_end}}
    )
    incident_count = await db.daily_logs.count_documents(
        {"company_id": cid, "submitted_at": {"$gte": week_start, "$lte": week_end}, "incidents": {"$nin": [None, ""]}}
    )
    photo_count = await db.project_photos.count_documents(
        {"company_id": cid, "uploaded_at": {"$gte": week_start, "$lte": week_end}}
    )
    summary = {
        "total_spend": total_spend,
        "log_count": log_count,
        "incident_count": incident_count,
        "photo_count": photo_count,
        "by_project": [{"project_id": s["_id"], "spend": round(s["spend"], 2)} for s in spend],
    }
    rep = {
        "id": str(uuid.uuid4()),
        "company_id": cid,
        "week_start": week_start,
        "week_end": week_end,
        "generated_at": now_utc(),
        "pdf_url": None,
        "excel_url": None,
        "summary": summary,
    }
    await db.weekly_reports.insert_one(rep.copy())
    return serialize(rep)


# =========================================================
# Profile / Company / GDPR / Forgot password
# =========================================================
@api.patch("/profile")
async def update_profile(body: ProfileUpdate, user: dict = Depends(get_current_user)):
    upd = {k: v for k, v in body.dict().items() if v is not None}
    if upd:
        upd["updated_at"] = now_utc()
        await db.users.update_one({"id": user["id"]}, {"$set": upd})
    u = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password_hash": 0})
    return serialize(u)


# ---- Notification preferences ----
@api.patch("/profile/notifications")
async def update_notification_preferences(
    body: NotificationPreferencesIn, user: dict = Depends(get_current_user)
):
    prefs = dict(user.get("notification_preferences") or DEFAULT_NOTIF_PREFS)
    incoming = body.dict(exclude_none=True)
    for k, v in incoming.items():
        if k in PREF_KEYS:
            prefs[k] = bool(v)
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"notification_preferences": prefs, "updated_at": now_utc()}},
    )
    return {"notification_preferences": prefs}


# ---- Push token registration ----
@api.post("/push-token")
async def add_push_token(body: PushTokenIn, user: dict = Depends(get_current_user)):
    return await register_push_token(
        db, user["id"], user["company_id"], body.token, body.platform or "unknown"
    )


@api.delete("/push-token")
async def remove_push_token(body: PushTokenDeleteIn, user: dict = Depends(get_current_user)):
    return await unregister_push_token(db, body.token)


@api.get("/company")
async def get_company(user: dict = Depends(get_current_user)):
    c = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0})
    return serialize(c) if c else {}


@api.patch("/company")
async def update_company(body: CompanyUpdate, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    upd = {k: v for k, v in body.dict().items() if v is not None}
    if upd:
        await db.companies.update_one({"id": user["company_id"]}, {"$set": upd})
    c = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0})
    return serialize(c)


@api.get("/gdpr/export")
async def gdpr_export(user: dict = Depends(get_current_user)):
    """Export all data for the user's company as JSON."""
    cid = user["company_id"]
    company = await db.companies.find_one({"id": cid}, {"_id": 0})
    users = await db.users.find({"company_id": cid}, {"_id": 0, "password_hash": 0}).to_list(1000)
    projects = await db.projects.find({"company_id": cid}, {"_id": 0}).to_list(1000)
    materials = await db.materials.find({"company_id": cid}, {"_id": 0}).to_list(2000)
    entries = await db.material_entries.find({"company_id": cid}, {"_id": 0}).to_list(5000)
    logs = await db.daily_logs.find({"company_id": cid}, {"_id": 0}).to_list(5000)
    photos = await db.project_photos.find(
        {"company_id": cid}, {"_id": 0, "image_base64": 0}  # exclude heavy field
    ).to_list(5000)
    alerts = await db.alerts.find({"company_id": cid}, {"_id": 0}).to_list(1000)
    payload = {
        "exported_at": iso(now_utc()),
        "company": serialize(company),
        "users": [serialize(u) for u in users],
        "projects": [serialize(p) for p in projects],
        "materials": [serialize(m) for m in materials],
        "material_entries": [serialize(e) for e in entries],
        "daily_logs": [serialize(l) for l in logs],
        "photos_metadata": [serialize(p) for p in photos],
        "alerts": [serialize(a) for a in alerts],
    }
    import json as _json, base64 as _b64
    raw = _json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    return {
        "filename": f"glasswork_export_{now_utc().strftime('%Y%m%d_%H%M%S')}.json",
        "mime": "application/json",
        "base64": _b64.b64encode(raw).decode("ascii"),
    }


@api.post("/auth/forgot-password")
async def forgot_password(body: ForgotPasswordIn):
    """Without email service: log a token to backend logs and return a generic message.
    Admin must contact user manually to share token. (No external email API used.)"""
    import secrets
    user = await db.users.find_one({"email": body.email.lower().strip()})
    # always return success to prevent enumeration
    if user:
        token = secrets.token_urlsafe(24)
        await db.password_reset_tokens.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "token": token,
            "expires_at": now_utc() + timedelta(hours=1),
            "used": False,
            "created_at": now_utc(),
        })
        logger.info(f"PASSWORD_RESET_TOKEN for {body.email}: {token}")
    return {"ok": True, "message": "Si el email existe, recibirás instrucciones por parte de tu administrador."}


# =========================================================
# Real PDF & Excel report generation
# =========================================================
def _build_pdf_bytes(company_name: str, summary: dict, week_start: datetime, week_end: datetime,
                    projects: list, logs: list, entries: list) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from io import BytesIO

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("title", parent=styles["Title"], fontSize=24, leading=28, textColor=colors.HexColor("#0A0A0A"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=14, textColor=colors.HexColor("#0A0A0A"))
    body = styles["BodyText"]
    cap = ParagraphStyle("cap", parent=body, fontSize=9, textColor=colors.HexColor("#525252"))

    story = []
    story.append(Paragraph(f"GLASSWORK", title))
    story.append(Paragraph(company_name, h2))
    story.append(Paragraph(f"Reporte semanal · {week_start.strftime('%d/%m/%Y')} – {week_end.strftime('%d/%m/%Y')}", cap))
    story.append(Spacer(1, 16))

    # KPIs
    story.append(Paragraph("Resumen ejecutivo", h2))
    kpi_data = [
        ["Gasto total", f"€{summary.get('total_spend', 0):,.2f}"],
        ["Partes registrados", str(summary.get('log_count', 0))],
        ["Fotos subidas", str(summary.get('photo_count', 0))],
        ["Incidentes", str(summary.get('incident_count', 0))],
    ]
    t = Table(kpi_data, colWidths=[80 * mm, 80 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F5F2")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0A0A0A")),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E5")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5E5")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))

    # Project breakdown
    story.append(Paragraph("Obras", h2))
    if projects:
        rows = [["Obra", "Estado", "Avance", "Presupuesto", "Gastado"]]
        for p in projects:
            rows.append([
                p.get("name", "—")[:38],
                p.get("status", "—"),
                f"{p.get('progress_percentage', 0)}%",
                f"€{p.get('budget', 0):,.0f}",
                f"€{p.get('spent', 0):,.0f}",
            ])
        pt = Table(rows, colWidths=[60 * mm, 24 * mm, 18 * mm, 30 * mm, 28 * mm])
        pt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 1), (-1, -1), "LEFT"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E5")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5E5")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAF8")]),
        ]))
        story.append(pt)
    else:
        story.append(Paragraph("Sin obras registradas en el período.", body))
    story.append(Spacer(1, 14))

    # Logs
    story.append(Paragraph("Partes diarios", h2))
    if logs:
        rows = [["Fecha", "Operario", "Obra", "Horas", "Estado"]]
        for l in logs[:30]:
            d = l.get("date", "")
            try:
                d_fmt = datetime.fromisoformat(d.replace("Z", "+00:00")).strftime("%d/%m/%Y") if d else "—"
            except Exception:
                d_fmt = d[:10] if d else "—"
            rows.append([
                d_fmt,
                (l.get("worker_name") or "—")[:24],
                (l.get("project_name") or "—")[:30],
                f"{l.get('hours_worked', 0)}h",
                l.get("status", "—"),
            ])
        lt = Table(rows, colWidths=[24 * mm, 36 * mm, 50 * mm, 18 * mm, 24 * mm])
        lt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E5")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5E5")),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAF8")]),
        ]))
        story.append(lt)
    else:
        story.append(Paragraph("Sin partes registrados.", body))

    story.append(Spacer(1, 14))
    story.append(Paragraph("Materiales", h2))
    if entries:
        cat_totals: dict = {}
        for e in entries:
            cat = (e.get("material") or {}).get("category", "OTROS")
            cat_totals[cat] = cat_totals.get(cat, 0) + (e.get("total_cost") or 0)
        rows = [["Categoría", "Total"]] + [[k, f"€{v:,.2f}"] for k, v in sorted(cat_totals.items(), key=lambda x: -x[1])]
        ct = Table(rows, colWidths=[60 * mm, 60 * mm])
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E5")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5E5")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(ct)

    doc.build(story)
    buf.seek(0)
    return buf.read()


def _build_excel_bytes(summary: dict, projects: list, logs: list, entries: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Métrica", "Valor"])
    for k, v in [
        ("Gasto total (€)", summary.get("total_spend", 0)),
        ("Partes", summary.get("log_count", 0)),
        ("Fotos", summary.get("photo_count", 0)),
        ("Incidentes", summary.get("incident_count", 0)),
    ]:
        ws.append([k, v])
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = ws["B1"].fill = PatternFill("solid", fgColor="0A0A0A")

    ws2 = wb.create_sheet("Obras")
    ws2.append(["Obra", "Estado", "Avance %", "Presupuesto", "Gastado", "Cliente"])
    for p in projects:
        ws2.append([p.get("name"), p.get("status"), p.get("progress_percentage", 0),
                    p.get("budget", 0), p.get("spent", 0), p.get("client_name", "")])
    for c in ws2[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="0A0A0A")

    ws3 = wb.create_sheet("Partes")
    ws3.append(["Fecha", "Operario", "Obra", "Horas", "Estado", "Descripción"])
    for l in logs:
        d = l.get("date", "")
        try:
            d_fmt = datetime.fromisoformat(d.replace("Z", "+00:00")).strftime("%Y-%m-%d") if d else ""
        except Exception:
            d_fmt = d[:10] if d else ""
        ws3.append([d_fmt, l.get("worker_name", ""), l.get("project_name", ""),
                    l.get("hours_worked", 0), l.get("status", ""), l.get("work_description", "")[:200]])
    for c in ws3[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="0A0A0A")

    ws4 = wb.create_sheet("Materiales")
    ws4.append(["Material", "Categoría", "Cantidad", "Unidad", "Precio", "Total", "Tipo", "Operario"])
    for e in entries:
        m = e.get("material") or {}
        ws4.append([m.get("name", ""), m.get("category", ""), e.get("quantity", 0),
                    m.get("unit", ""), e.get("unit_price", 0), e.get("total_cost", 0),
                    e.get("type", ""), e.get("worker_name", "")])
    for c in ws4[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="0A0A0A")

    for s in [ws, ws2, ws3, ws4]:
        for col in s.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            s.column_dimensions[col[0].column_letter].width = min(40, max_len + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def _gather_report_data(cid: str, week_start: datetime, week_end: datetime):
    projects = await db.projects.find({"company_id": cid}, {"_id": 0}).to_list(500)
    # add spent
    for p in projects:
        pipe = [{"$match": {"project_id": p["id"]}}, {"$group": {"_id": None, "s": {"$sum": "$total_cost"}}}]
        r = await db.material_entries.aggregate(pipe).to_list(1)
        p["spent"] = round(r[0]["s"], 2) if r else 0
    logs = await db.daily_logs.find(
        {"company_id": cid, "submitted_at": {"$gte": week_start, "$lte": week_end}}, {"_id": 0}
    ).sort("submitted_at", -1).to_list(500)
    for l in logs:
        w = await db.users.find_one({"id": l["worker_id"]}, {"_id": 0, "name": 1})
        proj = await db.projects.find_one({"id": l["project_id"]}, {"_id": 0, "name": 1})
        l["worker_name"] = (w or {}).get("name")
        l["project_name"] = (proj or {}).get("name")
    entries = await db.material_entries.find(
        {"company_id": cid, "created_at": {"$gte": week_start, "$lte": week_end}}, {"_id": 0}
    ).to_list(2000)
    for e in entries:
        m = await db.materials.find_one({"id": e["material_id"]}, {"_id": 0})
        e["material"] = serialize(m) if m else None
        w = await db.users.find_one({"id": e["worker_id"]}, {"_id": 0, "name": 1})
        e["worker_name"] = (w or {}).get("name")
    return projects, logs, entries


@api.get("/reports/weekly/{rid}/pdf")
async def report_pdf(rid: str, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    rep = await db.weekly_reports.find_one({"id": rid, "company_id": user["company_id"]})
    if not rep: raise HTTPException(status_code=404, detail="Reporte no encontrado")
    company = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0})
    ws = rep["week_start"] if isinstance(rep["week_start"], datetime) else datetime.fromisoformat(rep["week_start"])
    we = rep["week_end"] if isinstance(rep["week_end"], datetime) else datetime.fromisoformat(rep["week_end"])
    if ws.tzinfo is None: ws = ws.replace(tzinfo=timezone.utc)
    if we.tzinfo is None: we = we.replace(tzinfo=timezone.utc)
    projects, logs, entries = await _gather_report_data(user["company_id"], ws, we)
    pdf = _build_pdf_bytes(company.get("name", "Empresa"), rep.get("summary") or {}, ws, we, projects, logs, entries)
    import base64 as _b64
    return {
        "filename": f"glasswork_reporte_{ws.strftime('%Y%m%d')}.pdf",
        "mime": "application/pdf",
        "base64": _b64.b64encode(pdf).decode("ascii"),
    }


@api.get("/reports/weekly/{rid}/excel")
async def report_excel(rid: str, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    rep = await db.weekly_reports.find_one({"id": rid, "company_id": user["company_id"]})
    if not rep: raise HTTPException(status_code=404, detail="Reporte no encontrado")
    ws = rep["week_start"] if isinstance(rep["week_start"], datetime) else datetime.fromisoformat(rep["week_start"])
    we = rep["week_end"] if isinstance(rep["week_end"], datetime) else datetime.fromisoformat(rep["week_end"])
    if ws.tzinfo is None: ws = ws.replace(tzinfo=timezone.utc)
    if we.tzinfo is None: we = we.replace(tzinfo=timezone.utc)
    projects, logs, entries = await _gather_report_data(user["company_id"], ws, we)
    xls = _build_excel_bytes(rep.get("summary") or {}, projects, logs, entries)
    import base64 as _b64
    return {
        "filename": f"glasswork_reporte_{ws.strftime('%Y%m%d')}.xlsx",
        "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "base64": _b64.b64encode(xls).decode("ascii"),
    }


# =========================================================
# CLIENT REPORTS — one report per project (premium)
# =========================================================
@api.get("/reports/projects")
async def list_project_reports(
    status_f: Optional[str] = Query(None, alias="status"),
    user: dict = Depends(require_role("ADMIN", "MANAGER")),
):
    """List all projects with metrics so the UI can show report cards.
    Completed first, then active/paused/pending."""
    q: dict = {"company_id": user["company_id"]}
    if status_f:
        q["status"] = status_f
    projects = await db.projects.find(q).to_list(500)
    out = []
    status_order = {"COMPLETED": 0, "ACTIVE": 1, "PAUSED": 2, "PENDING": 3, "CANCELLED": 4}
    for p in projects:
        m = await project_metrics(p["id"])
        logs = await db.daily_logs.find({"project_id": p["id"]}).to_list(2000)
        hours_total = sum(float(l.get("hours_worked") or 0) for l in logs)
        worker_ids = list({l.get("worker_id") for l in logs if l.get("worker_id")})
        photos_count = m.get("photo_count", 0)
        out.append({
            "id": p["id"],
            "name": p.get("name"),
            "status": p.get("status"),
            "client_name": p.get("client_name", ""),
            "address": p.get("address", ""),
            "start_date": p.get("start_date"),
            "end_date": p.get("end_date"),
            "actual_end_date": p.get("actual_end_date"),
            "hours_total": round(hours_total, 1),
            "workers_count": len(worker_ids),
            "photo_count": photos_count,
            "log_count": m.get("log_count", 0),
            "progress_percentage": p.get("progress_percentage", 0),
            "cover_photo": p.get("cover_photo"),
        })
    out.sort(key=lambda x: (status_order.get(x["status"], 9), x.get("actual_end_date") or x.get("end_date") or ""), reverse=False)
    return out


@api.post("/projects/{project_id}/mark-complete")
async def mark_project_complete(
    project_id: str,
    user: dict = Depends(require_role("ADMIN", "MANAGER")),
):
    """Marks a project as COMPLETED setting actual_end_date to today."""
    p = await db.projects.find_one({"id": project_id, "company_id": user["company_id"]})
    if not p: raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    await db.projects.update_one(
        {"id": project_id, "company_id": user["company_id"]},
        {"$set": {
            "status": "COMPLETED",
            "actual_end_date": now_utc().isoformat(),
            "progress_percentage": 100,
            "updated_at": now_utc(),
        }}
    )
    p2 = await db.projects.find_one({"id": project_id})
    return serialize(p2)


async def _gather_project_report_data(project_id: str, company_id: str):
    """Collect all data needed for a premium client project report."""
    p = await db.projects.find_one({"id": project_id, "company_id": company_id})
    if not p: return None
    logs = await db.daily_logs.find({"project_id": project_id, "company_id": company_id}).sort("date", 1).to_list(5000)
    # attach worker names
    worker_ids = list({l.get("worker_id") for l in logs if l.get("worker_id")})
    workers_map: dict = {}
    if worker_ids:
        wusers = await db.users.find({"id": {"$in": worker_ids}}, {"_id": 0, "password_hash": 0}).to_list(500)
        workers_map = {w["id"]: w for w in wusers}
    entries = await db.material_entries.find({"project_id": project_id, "company_id": company_id}).to_list(5000)
    for e in entries:
        m = await db.materials.find_one({"id": e["material_id"]}, {"_id": 0})
        e["material"] = serialize(m) if m else None
    photos = await db.photos.find({"project_id": project_id, "company_id": company_id}).sort("created_at", 1).to_list(500)
    assigned_workers = await db.users.find(
        {"id": {"$in": p.get("assigned_worker_ids", [])}}, {"_id": 0, "password_hash": 0}
    ).to_list(100)
    return {
        "project": p,
        "logs": logs,
        "workers_map": workers_map,
        "entries": entries,
        "photos": photos,
        "assigned_workers": assigned_workers,
    }


def _build_client_project_pdf(company: dict, data: dict, manager_name: str) -> bytes:
    """Premium PDF for client handoff — no financial data."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame, Paragraph, Spacer, Table,
        TableStyle, PageBreak, Image as RLImage, KeepTogether,
    )
    from reportlab.pdfgen import canvas as _canvas
    from io import BytesIO
    import base64 as _b64

    buf = BytesIO()
    PAGE_W, PAGE_H = A4
    left = 18 * mm
    right = 18 * mm
    top = 20 * mm
    bottom = 22 * mm

    # ---------- Styles ----------
    styles = getSampleStyleSheet()
    DARK = colors.HexColor("#0A0A0A")
    GOLD = colors.HexColor("#B8924C")
    LIGHT = colors.HexColor("#F5F5F2")
    BORDER = colors.HexColor("#D4D4D4")
    MUTED = colors.HexColor("#525252")
    st_title = ParagraphStyle("tt", parent=styles["Title"], fontSize=30, leading=34, textColor=DARK, alignment=TA_LEFT, fontName="Helvetica-Bold")
    st_h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=18, leading=24, textColor=DARK, fontName="Helvetica-Bold", spaceAfter=6)
    st_h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, leading=18, textColor=DARK, fontName="Helvetica-Bold", spaceAfter=4)
    st_body = ParagraphStyle("b", parent=styles["BodyText"], fontSize=10, leading=14, textColor=DARK)
    st_cap = ParagraphStyle("cap", parent=styles["BodyText"], fontSize=8.5, leading=11, textColor=MUTED, fontName="Helvetica-Bold")
    st_cap.letterSpacing = 1.5
    st_cover_label = ParagraphStyle("cl", parent=styles["BodyText"], fontSize=9, leading=12, textColor=colors.white, fontName="Helvetica-Bold")
    st_cover_val = ParagraphStyle("cv", parent=styles["BodyText"], fontSize=13, leading=16, textColor=colors.white, fontName="Helvetica")
    st_center_muted = ParagraphStyle("cm", parent=styles["BodyText"], fontSize=10, leading=14, textColor=MUTED, alignment=TA_CENTER)

    # ---------- Document with custom footer ----------
    def draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(MUTED)
        footer = f"{company.get('name', 'GLASSWORK')} · Reporte de obra · Página {canvas.getPageNumber()}"
        canvas.drawString(left, 10 * mm, footer)
        canvas.drawRightString(PAGE_W - right, 10 * mm, now_utc().strftime("%d/%m/%Y"))
        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.3)
        canvas.line(left, 13 * mm, PAGE_W - right, 13 * mm)
        canvas.restoreState()

    frame = Frame(left, bottom, PAGE_W - left - right, PAGE_H - top - bottom, id="main", showBoundary=0)
    doc = BaseDocTemplate(buf, pagesize=A4, leftMargin=left, rightMargin=right, topMargin=top, bottomMargin=bottom)
    doc.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=draw_footer)])

    p = data["project"]
    story = []

    # ============================================
    # COVER (manual draw via canvas on first page)
    # ============================================
    def _fmt_date(s: Optional[str]) -> str:
        if not s: return "—"
        try:
            d = datetime.fromisoformat(s.replace("Z", "+00:00")) if isinstance(s, str) else s
            return d.strftime("%d / %m / %Y")
        except Exception:
            return str(s)[:10]

    # Build cover header block (dark band with GOLD accent line)
    from reportlab.platypus import Flowable
    class CoverHeader(Flowable):
        def __init__(self, w, h, company_name: str):
            super().__init__(); self.w = w; self.h = h; self.cname = company_name
        def wrap(self, aw, ah): return self.w, self.h
        def draw(self):
            c = self.canv
            c.setFillColor(DARK)
            c.rect(0, 0, self.w, self.h, fill=1, stroke=0)
            # Gold accent line
            c.setFillColor(GOLD)
            c.rect(0, self.h / 2 - 1, self.w, 2, fill=1, stroke=0)
            # Mono G
            box = 22
            c.setFillColor(colors.white)
            c.rect(14, self.h - 44, box, box, fill=1, stroke=0)
            c.setFillColor(DARK)
            c.setFont("Helvetica-Bold", 16)
            c.drawCentredString(14 + box / 2, self.h - 44 + 6, "G")
            # Company name
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 10)
            c.drawString(44, self.h - 30, "GLASSWORK")
            c.setFont("Helvetica", 9)
            c.drawString(44, self.h - 42, self.cname[:50])
            # Right-aligned label
            c.setFillColor(GOLD)
            c.setFont("Helvetica-Bold", 9)
            c.drawRightString(self.w - 12, self.h - 30, "REPORTE DE OBRA")
            c.setFillColor(colors.white)
            c.setFont("Helvetica", 8.5)
            c.drawRightString(self.w - 12, self.h - 42, "Certificado de obra finalizada")

    cw = PAGE_W - left - right
    story.append(CoverHeader(cw, 80, company.get("name", "GLASSWORK")))
    story.append(Spacer(1, 28))
    story.append(Paragraph(p.get("name", "Obra"), st_title))
    story.append(Spacer(1, 4))
    if p.get("address"):
        story.append(Paragraph(p.get("address"), st_body))
    story.append(Spacer(1, 24))

    # Client block (boxed)
    client_rows = [
        [Paragraph("CLIENTE", st_cap), Paragraph(p.get("client_name", "—") or "—", st_h2)],
    ]
    if p.get("client_phone"):
        client_rows.append([Paragraph("TELÉFONO", st_cap), Paragraph(p.get("client_phone"), st_body)])
    if p.get("client_email"):
        client_rows.append([Paragraph("EMAIL", st_cap), Paragraph(p.get("client_email"), st_body)])
    ct = Table(client_rows, colWidths=[28 * mm, cw - 28 * mm])
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBEFORE", (0, 0), (0, -1), 2, GOLD),
    ]))
    story.append(ct)
    story.append(Spacer(1, 16))

    # Dates block
    date_cells = [
        [Paragraph("INICIO DE OBRA", st_cap), Paragraph("FINALIZACIÓN", st_cap), Paragraph("ESTADO", st_cap)],
        [
            Paragraph(_fmt_date(p.get("start_date")), st_h2),
            Paragraph(_fmt_date(p.get("actual_end_date") or p.get("end_date")), st_h2),
            Paragraph("COMPLETADA" if p.get("status") == "COMPLETED" else "EN CURSO", st_h2),
        ],
    ]
    dt = Table(date_cells, colWidths=[cw / 3] * 3)
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
        ("BACKGROUND", (0, 1), (-1, 1), colors.white),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDER),
    ]))
    story.append(dt)

    story.append(PageBreak())

    # ============================================
    # PAGE 2: EXECUTIVE SUMMARY + DESCRIPTION
    # ============================================
    story.append(Paragraph("Resumen ejecutivo", st_h1))
    story.append(Paragraph("Datos globales de la obra realizada.", st_body))
    story.append(Spacer(1, 10))

    hours_total = sum(float(l.get("hours_worked") or 0) for l in data["logs"])
    worker_ids = list({l.get("worker_id") for l in data["logs"] if l.get("worker_id")})
    incidents = sum(1 for l in data["logs"] if (l.get("has_incident")))

    def _to_naive_date(v):
        """Robust parse: accepts datetime, ISO string with/without TZ, plain date string."""
        if v is None or v == "":
            return None
        if isinstance(v, datetime):
            return v.replace(tzinfo=None) if v.tzinfo else v
        if isinstance(v, str):
            try:
                d = datetime.fromisoformat(v.replace("Z", "+00:00"))
                return d.replace(tzinfo=None) if d.tzinfo else d
            except Exception:
                # Try plain date (YYYY-MM-DD)
                try:
                    return datetime.strptime(v[:10], "%Y-%m-%d")
                except Exception:
                    return None
        return None

    try:
        sd = _to_naive_date(p.get("start_date"))
        ed = _to_naive_date(p.get("actual_end_date") or p.get("end_date")) or now_utc().replace(tzinfo=None)
        days = max((ed - sd).days, 0) if sd else 0
    except Exception:
        days = 0

    kpi_cells = [
        [Paragraph("DURACIÓN", st_cap), Paragraph("HORAS TRABAJADAS", st_cap), Paragraph("OPERARIOS", st_cap), Paragraph("FOTOS", st_cap)],
        [
            Paragraph(f"{days}<font size=9> días</font>", st_h1),
            Paragraph(f"{hours_total:g}<font size=9> h</font>", st_h1),
            Paragraph(str(len(worker_ids)), st_h1),
            Paragraph(str(len(data["photos"])), st_h1),
        ],
    ]
    kt = Table(kpi_cells, colWidths=[cw / 4] * 4)
    kt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDER),
        ("BACKGROUND", (0, 1), (-1, 1), colors.white),
    ]))
    story.append(kt)
    story.append(Spacer(1, 16))

    if p.get("description"):
        story.append(Paragraph("Descripción", st_h2))
        story.append(Paragraph(p.get("description", "").replace("\n", "<br/>"), st_body))
        story.append(Spacer(1, 12))

    # ============================================
    # TEAM
    # ============================================
    if data["assigned_workers"] or worker_ids:
        story.append(Paragraph("Equipo que intervino", st_h2))
        # Build rows: name + hours
        hours_by = {}
        for l in data["logs"]:
            wid = l.get("worker_id")
            if wid: hours_by[wid] = hours_by.get(wid, 0) + float(l.get("hours_worked") or 0)
        # Include assigned workers even if 0 hours
        aw_ids = {w["id"] for w in data["assigned_workers"]}
        all_ids = aw_ids.union(hours_by.keys())
        name_map = {w["id"]: w.get("name", "—") for w in data["assigned_workers"]}
        name_map.update({wid: (data["workers_map"].get(wid) or {}).get("name", "—") for wid in hours_by.keys()})
        rows = [["Operario", "Horas totales"]]
        for wid in sorted(all_ids, key=lambda x: -hours_by.get(x, 0)):
            rows.append([name_map.get(wid, "—"), f"{hours_by.get(wid, 0):g} h"])
        tt = Table(rows, colWidths=[cw * 0.7, cw * 0.3])
        tt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAF8")]),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ]))
        story.append(tt)
        story.append(Spacer(1, 14))

    # ============================================
    # MATERIALS USED (no prices)
    # ============================================
    if data["entries"]:
        story.append(Paragraph("Materiales utilizados", st_h2))
        used_only = [e for e in data["entries"] if e.get("entry_type") in (None, "USAGE")]
        if not used_only: used_only = data["entries"]
        by_mat: dict = {}
        for e in used_only:
            mat = e.get("material") or {}
            key = mat.get("id") or mat.get("name", "?")
            if key not in by_mat:
                by_mat[key] = {"name": mat.get("name", "—"), "unit": mat.get("unit", ""), "category": mat.get("category", ""), "qty": 0.0}
            by_mat[key]["qty"] += float(e.get("quantity") or 0)
        rows = [["Material", "Categoría", "Cantidad"]]
        for v in sorted(by_mat.values(), key=lambda x: -x["qty"]):
            rows.append([v["name"][:60], v["category"], f"{v['qty']:g} {v['unit']}"])
        mt = Table(rows, colWidths=[cw * 0.5, cw * 0.25, cw * 0.25])
        mt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAF8")]),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ]))
        story.append(mt)

    # ============================================
    # PHOTO GALLERY (new page)
    # ============================================
    def _photo_flowable(b64: str, w: float, h: float):
        try:
            if not b64: return None
            raw = _b64.b64decode(b64)
            return RLImage(BytesIO(raw), width=w, height=h, kind="proportional")
        except Exception:
            return None

    photos_by_type = {}
    for ph in data["photos"]:
        t = ph.get("type") or "PROGRESS"
        photos_by_type.setdefault(t, []).append(ph)
    labels = [
        ("BEFORE", "Antes"),
        ("PROGRESS", "Durante"),
        ("AFTER", "Después"),
        ("INCIDENT", "Incidentes"),
        ("MATERIAL", "Material"),
        ("MEASUREMENT", "Medidas"),
    ]
    any_photo = any(photos_by_type.get(k) for k, _ in labels)
    if any_photo:
        story.append(PageBreak())
        story.append(Paragraph("Galería fotográfica", st_h1))
        story.append(Paragraph("Documentación visual de la obra.", st_body))
        story.append(Spacer(1, 10))
        cell_w = (cw - 8) / 2
        cell_h = cell_w * 0.72
        for key, label in labels:
            items = photos_by_type.get(key) or []
            if not items: continue
            story.append(Paragraph(label.upper(), st_cap))
            story.append(Spacer(1, 4))
            # Render 2 columns
            pairs = []
            row: list = []
            for ph in items[:12]:
                imgf = _photo_flowable(ph.get("base64") or ph.get("data", ""), cell_w, cell_h)
                row.append(imgf if imgf else Paragraph("(imagen no disponible)", st_center_muted))
                if len(row) == 2:
                    pairs.append(row); row = []
            if row:
                while len(row) < 2: row.append("")
                pairs.append(row)
            if pairs:
                gt = Table(pairs, colWidths=[cell_w, cell_w], rowHeights=[cell_h] * len(pairs))
                gt.setStyle(TableStyle([
                    ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]))
                story.append(gt)
            story.append(Spacer(1, 8))

    # ============================================
    # SIGNATURE
    # ============================================
    story.append(Spacer(1, 22))
    story.append(Paragraph("Certificación", st_h1))
    story.append(Paragraph(
        f"Certificamos que la obra <b>{p.get('name', '')}</b> ubicada en "
        f"<b>{p.get('address', '—')}</b> ha sido ejecutada y finalizada conforme a lo acordado con el cliente "
        f"<b>{p.get('client_name', '—')}</b>.",
        st_body,
    ))
    story.append(Spacer(1, 18))
    sign_cells = [
        [Paragraph("RESPONSABLE", st_cap), Paragraph("FECHA DE CERTIFICACIÓN", st_cap)],
        [Paragraph(manager_name or "—", st_h2), Paragraph(now_utc().strftime("%d / %m / %Y"), st_h2)],
        [Paragraph(f"<i>{company.get('name', 'GLASSWORK')}</i>", st_body), Paragraph("<i>Firma digital verificada</i>", st_body)],
    ]
    sg = Table(sign_cells, colWidths=[cw / 2, cw / 2])
    sg.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
        ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDER),
        ("LINEBEFORE", (0, 0), (0, -1), 2, GOLD),
    ]))
    story.append(sg)

    doc.build(story)
    buf.seek(0)
    return buf.read()


@api.get("/projects/{project_id}/client-report/pdf")
async def client_project_report_pdf(
    project_id: str,
    user: dict = Depends(require_role("ADMIN", "MANAGER")),
):
    """Generate a premium PDF report of a project to share with the client. No financial data."""
    company = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0}) or {}
    data = await _gather_project_report_data(project_id, user["company_id"])
    if not data:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    manager_name = user.get("name", "")
    pdf = _build_client_project_pdf(company, data, manager_name)
    import base64 as _b64
    proj = data["project"]
    safe_name = "".join(c for c in proj.get("name", "obra") if c.isalnum() or c in (" ", "-", "_")).strip().replace(" ", "_") or "obra"
    return {
        "filename": f"glasswork_reporte_{safe_name}_{now_utc().strftime('%Y%m%d')}.pdf",
        "mime": "application/pdf",
        "base64": _b64.b64encode(pdf).decode("ascii"),
    }



# =========================================================
# WAREHOUSE — lots, zones, movements, label printing
# =========================================================
LotStatus = Literal["IN_STOCK", "PARTIAL", "DEPLETED"]
MovementType = Literal["INBOUND", "OUTBOUND", "LOCATE", "ADJUST", "RETURN"]


class StorageZoneIn(BaseModel):
    name: str
    category: MatCategory
    row_count: int = 10


# ---- New: physical storage locations (zone × row × material) ----
class WarehouseImportItem(BaseModel):
    """A single physical slot inside a zone."""
    zone_number: int       # 1..N
    row_number: int        # 1..row_count
    material_code: str
    quantity: float = 0
    min_quantity: float = 5


class WarehouseImportZone(BaseModel):
    zone_number: int
    name: str
    category: MatCategory = "PERFILERIA"
    row_count: int = 12


class WarehouseImportMaterial(BaseModel):
    code: str
    name: str
    category: MatCategory = "PERFILERIA"
    unit: str = "m"               # 'm' / 'ud' / 'kg' …
    supplier: str = "Cortizo"
    family: Optional[str] = None  # e.g. "COR 70 INDUSTRIAL", "COR VISION EVOLUTION"


class WarehouseImportPayload(BaseModel):
    """Single transactional import for the whole warehouse planning.

    The endpoint is idempotent: zones / materials are upserted by name/code,
    and locations are replaced for this company.
    """
    materials: List[WarehouseImportMaterial] = []
    zones: List[WarehouseImportZone] = []
    locations: List[WarehouseImportItem] = []
    wipe_existing_materials: bool = False  # ⚠ deletes all materials of the company first


class LocationStockIn(BaseModel):
    """In/out stock movement for a physical location (tablet flow)."""
    delta: float                          # positive = inbound, negative = outbound
    project_id: Optional[str] = None      # optional reference for outbound
    note: Optional[str] = ""


class LotCreateIn(BaseModel):
    material_id: str
    quantity: float
    supplier_name: Optional[str] = ""
    unit_price: Optional[float] = None
    notes: Optional[str] = ""


class LocateIn(BaseModel):
    zone_id: str
    row_label: Optional[str] = ""


class OutboundIn(BaseModel):
    quantity: float
    project_id: str
    note: Optional[str] = ""


class AdjustIn(BaseModel):
    quantity: float  # can be negative
    note: str = ""


async def _next_lot_code() -> str:
    year = now_utc().year
    prefix = f"EG-{year}-"
    last = await db.material_lots.find({"lot_code": {"$regex": f"^{prefix}"}}).sort("lot_code", -1).limit(1).to_list(1)
    if not last:
        n = 1
    else:
        try: n = int(last[0]["lot_code"].split("-")[-1]) + 1
        except Exception: n = 1
    return f"{prefix}{n:04d}"


def _lot_status(qty_left: float, qty_total: float) -> str:
    if qty_left <= 0: return "DEPLETED"
    if qty_left < qty_total: return "PARTIAL"
    return "IN_STOCK"


@api.post("/warehouse/zones")
async def warehouse_create_zone(body: StorageZoneIn, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    zid = str(uuid.uuid4())
    z = {
        "id": zid,
        "company_id": user["company_id"],
        "name": body.name,
        "category": body.category,
        "row_count": body.row_count,
        "qr_code": f"ZONE-{zid}",
        "created_at": now_utc(),
    }
    await db.storage_zones.insert_one(z.copy())
    return serialize(z)


@api.get("/warehouse/zones")
async def warehouse_list_zones(user: dict = Depends(get_current_user)):
    zones = await db.storage_zones.find({"company_id": user["company_id"]}).sort("name", 1).to_list(200)
    out = []
    for z in zones:
        z = serialize(z)
        z["lot_count"] = await db.material_lots.count_documents({"company_id": user["company_id"], "zone_id": z["id"]})
        out.append(z)
    return out


# ---- New: physical storage locations (warehouse map) ----

def _loc_status(quantity: float, min_q: float) -> str:
    """Return color code based on stock vs threshold."""
    if quantity <= 0:
        return "OUT"        # red
    if quantity <= max(min_q, 0):
        return "LOW"        # yellow
    return "OK"             # green


@api.post("/warehouse/import-locations")
async def warehouse_import_locations(
    body: WarehouseImportPayload,
    user: dict = Depends(require_role("ADMIN", "MANAGER")),
):
    """Idempotent bulk import of the whole warehouse planning.

    - Optionally wipes existing materials (`wipe_existing_materials`).
    - Upserts the material catalog (by `code`).
    - Upserts zones (by `zone_number`).
    - Replaces all `storage_locations` for this company.
    - Generates QR codes Z{n}-F{row}-{code} for every location.
    """
    company_id = user["company_id"]
    summary = {"materials": 0, "zones": 0, "locations": 0, "wiped": 0, "qrs": 0}

    # 1) Optional: wipe existing materials of the company
    if body.wipe_existing_materials:
        # Delete materials + dependent lots/locations
        del_mats = await db.materials.delete_many({"company_id": company_id})
        del_locs = await db.storage_locations.delete_many({"company_id": company_id}) \
            if "storage_locations" in await db.list_collection_names() else None
        await db.material_lots.delete_many({"company_id": company_id})
        summary["wiped"] = int(del_mats.deleted_count or 0)

    # 2) Materials — upsert by (company_id, code)
    code_to_id: dict[str, str] = {}
    for m in body.materials:
        existing = await db.materials.find_one({"company_id": company_id, "code": m.code}, {"_id": 0, "id": 1})
        if existing:
            mid = existing["id"]
            await db.materials.update_one(
                {"id": mid},
                {"$set": {
                    "name": m.name,
                    "category": m.category,
                    "unit": m.unit,
                    "supplier": m.supplier,
                    "family": m.family or "",
                    "updated_at": now_utc(),
                }},
            )
        else:
            mid = str(uuid.uuid4())
            await db.materials.insert_one({
                "id": mid,
                "company_id": company_id,
                "code": m.code,
                "name": m.name,
                "category": m.category,
                "unit": m.unit,
                "supplier": m.supplier,
                "family": m.family or "",
                "unit_price": 0,
                "stock": 0,
                "min_stock": 0,
                "is_active": True,
                "created_at": now_utc(),
                "updated_at": now_utc(),
            })
        code_to_id[m.code] = mid
        summary["materials"] += 1

    # Build a code → id map also from existing materials (so locations can reference them)
    if not code_to_id:
        existing = await db.materials.find(
            {"company_id": company_id}, {"_id": 0, "id": 1, "code": 1}
        ).to_list(5000)
        code_to_id = {e["code"]: e["id"] for e in existing}

    # 3) Zones — upsert by (company_id, zone_number)
    zone_num_to_id: dict[int, dict] = {}
    for z in body.zones:
        existing = await db.storage_zones.find_one(
            {"company_id": company_id, "zone_number": z.zone_number}, {"_id": 0}
        )
        if existing:
            zid = existing["id"]
            await db.storage_zones.update_one(
                {"id": zid},
                {"$set": {
                    "name": z.name,
                    "category": z.category,
                    "row_count": z.row_count,
                    "updated_at": now_utc(),
                }},
            )
        else:
            zid = str(uuid.uuid4())
            await db.storage_zones.insert_one({
                "id": zid,
                "company_id": company_id,
                "zone_number": z.zone_number,
                "name": z.name,
                "category": z.category,
                "row_count": z.row_count,
                "qr_code": f"GW-ZONE-{zid[:8].upper()}",
                "created_at": now_utc(),
                "updated_at": now_utc(),
            })
        zone_num_to_id[z.zone_number] = {"id": zid, "name": z.name}
        summary["zones"] += 1

    # 4) Locations — replace all for the company
    await db.storage_locations.delete_many({"company_id": company_id})

    locs_to_insert = []
    for loc in body.locations:
        zone = zone_num_to_id.get(loc.zone_number)
        if not zone:
            # zone not provided in payload — try to look it up
            z_existing = await db.storage_zones.find_one(
                {"company_id": company_id, "zone_number": loc.zone_number}, {"_id": 0, "id": 1, "name": 1}
            )
            if not z_existing:
                continue
            zone = {"id": z_existing["id"], "name": z_existing["name"]}
        material_id = code_to_id.get(loc.material_code)
        if not material_id:
            # Material not found — skip silently
            continue
        qr = f"Z{loc.zone_number}-F{loc.row_number}-{loc.material_code}"
        locs_to_insert.append({
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "zone_id": zone["id"],
            "zone_number": loc.zone_number,
            "zone_name": zone["name"],
            "row_number": loc.row_number,
            "material_id": material_id,
            "material_code": loc.material_code,
            "quantity": float(loc.quantity or 0),
            "min_quantity": float(loc.min_quantity or 0),
            "qr_code": qr,
            "status": _loc_status(loc.quantity or 0, loc.min_quantity or 0),
            "created_at": now_utc(),
            "updated_at": now_utc(),
        })
        summary["qrs"] += 1
    if locs_to_insert:
        await db.storage_locations.insert_many(locs_to_insert)
        summary["locations"] = len(locs_to_insert)

    await audit_log(
        db, action="WAREHOUSE_IMPORT", resource="warehouse",
        request=None, user=user, success=True, extra=summary,
    )
    return summary


@api.get("/warehouse/locations")
async def warehouse_list_locations(
    zone_number: Optional[int] = None,
    user: dict = Depends(get_current_user),
):
    flt: dict = {"company_id": user["company_id"]}
    if zone_number is not None:
        flt["zone_number"] = zone_number
    locs = await db.storage_locations.find(flt).sort([("zone_number", 1), ("row_number", 1)]).to_list(2000)
    out = []
    for loc in locs:
        loc = serialize(loc)
        loc["status"] = _loc_status(loc.get("quantity", 0), loc.get("min_quantity", 0))
        # Attach material display info
        mat = await db.materials.find_one({"id": loc["material_id"]}, {"_id": 0, "name": 1, "unit": 1, "category": 1, "family": 1, "supplier": 1, "code": 1})
        loc["material"] = mat or {"name": "", "unit": "", "code": loc.get("material_code", "")}
        out.append(loc)
    return out


@api.get("/warehouse/locations/by-qr/{qr_code}")
async def warehouse_location_by_qr(qr_code: str, user: dict = Depends(get_current_user)):
    loc = await db.storage_locations.find_one(
        {"company_id": user["company_id"], "qr_code": qr_code}, {"_id": 0}
    )
    if not loc:
        raise HTTPException(status_code=404, detail="Ubicación no encontrada")
    loc["status"] = _loc_status(loc.get("quantity", 0), loc.get("min_quantity", 0))
    mat = await db.materials.find_one({"id": loc["material_id"]}, {"_id": 0, "name": 1, "unit": 1, "category": 1, "supplier": 1, "code": 1, "family": 1})
    loc["material"] = mat or {}
    return loc


@api.post("/warehouse/locations/{loc_id}/stock")
async def warehouse_location_stock(
    loc_id: str,
    body: LocationStockIn,
    user: dict = Depends(get_current_user),
):
    """Single-step stock movement on a physical location (tablet flow)."""
    if body.delta == 0:
        raise HTTPException(status_code=400, detail="Cantidad debe ser distinta de 0")
    loc = await db.storage_locations.find_one(
        {"id": loc_id, "company_id": user["company_id"]}, {"_id": 0}
    )
    if not loc:
        raise HTTPException(status_code=404, detail="Ubicación no encontrada")
    new_qty = float(loc.get("quantity", 0)) + float(body.delta)
    if new_qty < 0:
        raise HTTPException(status_code=400, detail=f"Stock insuficiente. Quedan {loc.get('quantity', 0)}.")
    await db.storage_locations.update_one(
        {"id": loc_id},
        {"$set": {
            "quantity": new_qty,
            "status": _loc_status(new_qty, loc.get("min_quantity", 0)),
            "updated_at": now_utc(),
        }},
    )
    # Log movement
    await db.lot_movements.insert_one({
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "location_id": loc_id,
        "material_id": loc["material_id"],
        "material_code": loc.get("material_code", ""),
        "qr_code": loc.get("qr_code", ""),
        "type": "INBOUND" if body.delta > 0 else "OUTBOUND",
        "quantity": abs(float(body.delta)),
        "project_id": body.project_id or None,
        "note": body.note or "",
        "user_id": user["id"],
        "user_name": user.get("name", ""),
        "created_at": now_utc(),
    })
    loc["quantity"] = new_qty
    loc["status"] = _loc_status(new_qty, loc.get("min_quantity", 0))
    return loc


@api.get("/warehouse/zones/{zid}/qr.png")
async def warehouse_zone_qr(zid: str, user: dict = Depends(get_current_user)):
    from fastapi.responses import Response
    z = await db.storage_zones.find_one({"id": zid, "company_id": user["company_id"]})
    if not z: raise HTTPException(status_code=404, detail="Zona no encontrada")
    import qrcode
    from io import BytesIO
    img = qrcode.make(z["qr_code"], box_size=10, border=2)
    buf = BytesIO(); img.save(buf, format="PNG")
    return Response(content=buf.getvalue(), media_type="image/png")


@api.get("/warehouse/zones/{zid}/qr-base64")
async def warehouse_zone_qr_b64(zid: str, user: dict = Depends(get_current_user)):
    """Return zone QR code as base64 PNG (works with auth in mobile/web)."""
    z = await db.storage_zones.find_one({"id": zid, "company_id": user["company_id"]})
    if not z: raise HTTPException(status_code=404, detail="Zona no encontrada")
    import qrcode, base64
    from io import BytesIO
    img = qrcode.make(z["qr_code"], box_size=10, border=2)
    buf = BytesIO(); img.save(buf, format="PNG")
    return {
        "filename": f"zone-{z.get('name', zid).replace(' ', '_')}.png",
        "mime": "image/png",
        "base64": base64.b64encode(buf.getvalue()).decode("ascii"),
        "qr_code": z["qr_code"],
    }


@api.post("/warehouse/lots")
async def warehouse_create_lot(request: Request, body: LotCreateIn, user: dict = Depends(get_current_user)):
    mat = await db.materials.find_one({"id": body.material_id, "company_id": user["company_id"]})
    if not mat: raise HTTPException(status_code=404, detail="Material no encontrado")
    code = await _next_lot_code()
    unit_price = body.unit_price if body.unit_price is not None else mat.get("unit_price", 0)
    lot = {
        "id": str(uuid.uuid4()),
        "lot_code": code,
        "company_id": user["company_id"],
        "material_id": body.material_id,
        "quantity": body.quantity,
        "quantity_left": body.quantity,
        "supplier_name": body.supplier_name or mat.get("supplier", ""),
        "unit_price": unit_price,
        "entry_date": now_utc(),
        "registered_by": user["id"],
        "zone_id": None,
        "row_label": None,
        "status": "IN_STOCK",
        "notes": body.notes or "",
        "created_at": now_utc(),
    }
    await db.material_lots.insert_one(lot.copy())
    await db.lot_movements.insert_one({
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "lot_id": lot["id"],
        "lot_code": code,
        "type": "INBOUND",
        "quantity": body.quantity,
        "project_id": None,
        "worker_id": user["id"],
        "timestamp": now_utc(),
        "note": body.notes or "",
    })
    out = serialize(lot)
    out["material"] = serialize(mat)
    if user["role"] != "ADMIN":
        out["unit_price"] = None
    await audit_log(
        db, action="WAREHOUSE_MOVE_INBOUND", resource="lot", resource_id=lot["id"],
        request=request, user=user,
        extra={"lot_code": code, "quantity": body.quantity, "material_id": body.material_id},
    )
    return out


@api.get("/warehouse/lots")
async def warehouse_list_lots(
    zone_id: Optional[str] = None,
    status_f: Optional[str] = Query(None, alias="status"),
    material_id: Optional[str] = None,
    q: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    flt: dict = {"company_id": user["company_id"]}
    if zone_id: flt["zone_id"] = zone_id
    if status_f: flt["status"] = status_f
    if material_id: flt["material_id"] = material_id
    if q: flt["lot_code"] = {"$regex": q, "$options": "i"}
    lots = await db.material_lots.find(flt).sort("entry_date", -1).to_list(500)
    out = []
    for l in lots:
        l = serialize(l)
        m = await db.materials.find_one({"id": l["material_id"]}, {"_id": 0})
        l["material"] = serialize(m) if m else None
        if user["role"] == "WORKER":
            l["unit_price"] = None
        out.append(l)
    return out


@api.get("/warehouse/lots/{lot_code}")
async def warehouse_lot_detail(lot_code: str, user: dict = Depends(get_current_user)):
    lot = await db.material_lots.find_one({"lot_code": lot_code, "company_id": user["company_id"]})
    if not lot: raise HTTPException(status_code=404, detail="Lote no encontrado")
    lot = serialize(lot)
    mat = await db.materials.find_one({"id": lot["material_id"]}, {"_id": 0})
    lot["material"] = serialize(mat) if mat else None
    if lot.get("zone_id"):
        z = await db.storage_zones.find_one({"id": lot["zone_id"]}, {"_id": 0})
        lot["zone"] = serialize(z) if z else None
    movs = await db.lot_movements.find({"lot_id": lot["id"]}).sort("timestamp", -1).to_list(100)
    enriched = []
    for m in movs:
        m = serialize(m)
        if m.get("project_id"):
            p = await db.projects.find_one({"id": m["project_id"]}, {"_id": 0, "name": 1})
            m["project_name"] = (p or {}).get("name")
        u = await db.users.find_one({"id": m["worker_id"]}, {"_id": 0, "name": 1})
        m["worker_name"] = (u or {}).get("name")
        enriched.append(m)
    lot["movements"] = enriched
    if user["role"] != "ADMIN":
        lot["unit_price"] = None
    return lot


@api.post("/warehouse/lots/{lot_code}/locate")
async def warehouse_locate(request: Request, lot_code: str, body: LocateIn, user: dict = Depends(get_current_user)):
    lot = await db.material_lots.find_one({"lot_code": lot_code, "company_id": user["company_id"]})
    if not lot: raise HTTPException(status_code=404, detail="Lote no encontrado")
    z = await db.storage_zones.find_one({"id": body.zone_id, "company_id": user["company_id"]})
    if not z: raise HTTPException(status_code=404, detail="Zona no encontrada")
    await db.material_lots.update_one({"id": lot["id"]}, {"$set": {"zone_id": body.zone_id, "row_label": body.row_label or "", "updated_at": now_utc()}})
    await db.lot_movements.insert_one({
        "id": str(uuid.uuid4()), "company_id": user["company_id"], "lot_id": lot["id"], "lot_code": lot_code,
        "type": "LOCATE", "quantity": 0, "project_id": None, "worker_id": user["id"],
        "timestamp": now_utc(), "note": f"Zona {z['name']} {body.row_label or ''}".strip(),
    })
    await audit_log(
        db, action="WAREHOUSE_MOVE_LOCATE", resource="lot", resource_id=lot["id"],
        request=request, user=user,
        extra={"lot_code": lot_code, "zone": z.get("name"), "row_label": body.row_label or ""},
    )
    return {"ok": True, "zone": z["name"], "row": body.row_label or ""}


@api.post("/warehouse/lots/{lot_code}/outbound")
async def warehouse_outbound(request: Request, lot_code: str, body: OutboundIn, user: dict = Depends(get_current_user)):
    lot = await db.material_lots.find_one({"lot_code": lot_code, "company_id": user["company_id"]})
    if not lot: raise HTTPException(status_code=404, detail="Lote no encontrado")
    if body.quantity <= 0: raise HTTPException(status_code=400, detail="Cantidad debe ser positiva")
    if body.quantity > lot["quantity_left"]:
        raise HTTPException(status_code=409, detail=f"Stock insuficiente. Disponible: {lot['quantity_left']}")
    proj = await db.projects.find_one({"id": body.project_id, "company_id": user["company_id"]})
    if not proj: raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if user["role"] == "WORKER" and user["id"] not in proj.get("assigned_worker_ids", []):
        raise HTTPException(status_code=403, detail="Sin acceso al proyecto")
    new_left = lot["quantity_left"] - body.quantity
    new_status = _lot_status(new_left, lot["quantity"])
    await db.material_lots.update_one({"id": lot["id"]}, {"$set": {"quantity_left": new_left, "status": new_status, "updated_at": now_utc()}})
    await db.lot_movements.insert_one({
        "id": str(uuid.uuid4()), "company_id": user["company_id"], "lot_id": lot["id"], "lot_code": lot_code,
        "type": "OUTBOUND", "quantity": body.quantity, "project_id": body.project_id, "worker_id": user["id"],
        "timestamp": now_utc(), "note": body.note or "",
    })
    # also create MaterialEntry on the project (USAGE) so existing project finance flows pick it up
    unit_price = lot.get("unit_price") or 0
    await db.material_entries.insert_one({
        "id": str(uuid.uuid4()), "company_id": user["company_id"],
        "project_id": body.project_id, "material_id": lot["material_id"],
        "quantity": body.quantity, "unit_price": unit_price, "total_cost": round(unit_price * body.quantity, 2),
        "type": "USAGE", "worker_id": user["id"], "date": now_utc().isoformat(),
        "notes": f"Salida lote {lot_code}", "receipt_photo": None, "created_at": now_utc(),
    })
    await audit_log(
        db, action="WAREHOUSE_MOVE_OUTBOUND", resource="lot", resource_id=lot["id"],
        request=request, user=user,
        extra={"lot_code": lot_code, "quantity": body.quantity, "project_id": body.project_id},
    )
    return {"ok": True, "quantity_left": new_left, "status": new_status}


@api.post("/warehouse/lots/{lot_code}/adjust")
async def warehouse_adjust(request: Request, lot_code: str, body: AdjustIn, user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    lot = await db.material_lots.find_one({"lot_code": lot_code, "company_id": user["company_id"]})
    if not lot: raise HTTPException(status_code=404, detail="Lote no encontrado")
    new_left = max(0.0, lot["quantity_left"] + body.quantity)
    new_status = _lot_status(new_left, lot["quantity"])
    await db.material_lots.update_one({"id": lot["id"]}, {"$set": {"quantity_left": new_left, "status": new_status, "updated_at": now_utc()}})
    await db.lot_movements.insert_one({
        "id": str(uuid.uuid4()), "company_id": user["company_id"], "lot_id": lot["id"], "lot_code": lot_code,
        "type": "ADJUST", "quantity": body.quantity, "project_id": None, "worker_id": user["id"],
        "timestamp": now_utc(), "note": body.note or "",
    })
    return {"ok": True, "quantity_left": new_left, "status": new_status}


@api.get("/warehouse/stock")
async def warehouse_stock(user: dict = Depends(get_current_user)):
    """Aggregated stock per material — combines lots (legacy) AND storage_locations (map flow).

    The map UI updates `storage_locations.quantity` directly so its data must be reflected
    here as the primary source. We additionally include any non-depleted `material_lots`
    rows to remain backwards compatible with the older receive flow.
    """
    cid = user["company_id"]
    totals: dict = {}  # material_id -> {total, lot_count (=positions), min_total}

    # 1) Primary: storage_locations (where the map flow actually writes)
    locs_pipe = [
        {"$match": {"company_id": cid}},
        {"$group": {
            "_id": "$material_id",
            "total": {"$sum": "$quantity"},
            "lot_count": {"$sum": 1},
            "min_total": {"$sum": "$min_quantity"},
        }},
    ]
    for r in await db.storage_locations.aggregate(locs_pipe).to_list(5000):
        totals[r["_id"]] = {
            "total": float(r["total"] or 0),
            "lot_count": int(r["lot_count"] or 0),
            "min_total": float(r["min_total"] or 0),
            "value": 0.0,
        }

    # 2) Legacy: material_lots (if any non-depleted records exist)
    lots_pipe = [
        {"$match": {"company_id": cid, "status": {"$ne": "DEPLETED"}}},
        {"$group": {
            "_id": "$material_id",
            "total": {"$sum": "$quantity_left"},
            "lot_count": {"$sum": 1},
            "value": {"$sum": {"$multiply": ["$quantity_left", "$unit_price"]}},
        }},
    ]
    for r in await db.material_lots.aggregate(lots_pipe).to_list(2000):
        cur = totals.setdefault(r["_id"], {"total": 0.0, "lot_count": 0, "min_total": 0.0, "value": 0.0})
        cur["total"] += float(r["total"] or 0)
        cur["lot_count"] += int(r["lot_count"] or 0)
        cur["value"] += float(r["value"] or 0)

    if not totals:
        return []

    # 3) Hydrate materials in one query
    mids = list(totals.keys())
    mats = await db.materials.find({"id": {"$in": mids}}, {"_id": 0}).to_list(len(mids))
    mat_by_id = {m["id"]: m for m in mats}

    out = []
    for mid, agg in totals.items():
        m = mat_by_id.get(mid)
        if not m:
            continue
        # Low stock: either total ≤ aggregated min (if defined) or below default threshold
        threshold = agg["min_total"] if agg["min_total"] > 0 else 5
        item = {
            "material_id": mid,
            "name": m.get("name") or m.get("code") or "—",
            "category": m.get("category") or "OTROS",
            "unit": m.get("unit") or "u",
            "total": round(agg["total"], 2),
            "lot_count": agg["lot_count"],  # number of physical positions/lots
            "low_stock": agg["total"] <= threshold and agg["total"] >= 0,
        }
        if user["role"] == "ADMIN":
            item["value"] = round(agg["value"], 2)
        out.append(item)
    out.sort(key=lambda x: (x["category"], x["name"]))
    return out


@api.get("/warehouse/zones/by-qr/{qr_code}")
async def warehouse_zone_by_qr(qr_code: str, user: dict = Depends(get_current_user)):
    """Resolve a zone QR (GW-ZONE-XXXXXXXX) to its zone + all locations inside.

    Used by the mobile scanner: scan a zone label and immediately see all
    materials stored in that zone with stock for quick IN/OUT.
    """
    z = await db.storage_zones.find_one(
        {"company_id": user["company_id"], "qr_code": qr_code}, {"_id": 0}
    )
    if not z:
        raise HTTPException(status_code=404, detail="Zona no encontrada")
    locs = await db.storage_locations.find(
        {"company_id": user["company_id"], "zone_id": z["id"]}, {"_id": 0}
    ).sort("row_number", 1).to_list(500)
    # Hydrate material info
    mids = list({l["material_id"] for l in locs if l.get("material_id")})
    mats = await db.materials.find({"id": {"$in": mids}}, {"_id": 0}).to_list(len(mids)) if mids else []
    mat_by_id = {m["id"]: m for m in mats}
    for l in locs:
        l["status"] = _loc_status(l.get("quantity", 0), l.get("min_quantity", 0))
        l["material"] = mat_by_id.get(l.get("material_id")) or {"name": "", "unit": "", "code": l.get("material_code", "")}
    return {"zone": z, "locations": locs}


@api.get("/warehouse/movements")
async def warehouse_movements(
    project_id: Optional[str] = None,
    worker_id: Optional[str] = None,
    type_f: Optional[str] = Query(None, alias="type"),
    user: dict = Depends(get_current_user),
):
    flt: dict = {"company_id": user["company_id"]}
    if project_id: flt["project_id"] = project_id
    if worker_id: flt["worker_id"] = worker_id
    if type_f: flt["type"] = type_f
    if user["role"] == "WORKER": flt["worker_id"] = user["id"]
    movs = await db.lot_movements.find(flt).sort("timestamp", -1).to_list(500)
    out = []
    for m in movs:
        m = serialize(m)
        u = await db.users.find_one({"id": m["worker_id"]}, {"_id": 0, "name": 1})
        m["worker_name"] = (u or {}).get("name")
        if m.get("project_id"):
            p = await db.projects.find_one({"id": m["project_id"]}, {"_id": 0, "name": 1})
            m["project_name"] = (p or {}).get("name")
        out.append(m)
    return out


@api.get("/warehouse/dashboard")
async def warehouse_dashboard(user: dict = Depends(require_role("ADMIN", "MANAGER"))):
    cid = user["company_id"]
    today = datetime(now_utc().year, now_utc().month, now_utc().day, tzinfo=timezone.utc)

    # Active storage positions (replaces "lots_count" — the map flow uses storage_locations).
    # Legacy lots are added on top for backwards compatibility.
    positions_count = await db.storage_locations.count_documents(
        {"company_id": cid, "quantity": {"$gt": 0}}
    )
    legacy_lots_count = await db.material_lots.count_documents(
        {"company_id": cid, "status": {"$ne": "DEPLETED"}}
    )
    lots_count = positions_count + legacy_lots_count

    movements_today = await db.lot_movements.count_documents({"company_id": cid, "timestamp": {"$gte": today}})

    # Stock value: storage_locations.quantity × materials.unit_price (via $lookup) + legacy lots
    loc_value_pipe = [
        {"$match": {"company_id": cid, "quantity": {"$gt": 0}}},
        {"$lookup": {
            "from": "materials", "localField": "material_id", "foreignField": "id",
            "as": "mat",
        }},
        {"$unwind": {"path": "$mat", "preserveNullAndEmptyArrays": True}},
        {"$group": {"_id": None, "v": {"$sum": {"$multiply": ["$quantity", {"$ifNull": ["$mat.unit_price", 0]}]}}}},
    ]
    loc_val = await db.storage_locations.aggregate(loc_value_pipe).to_list(1)
    legacy_val_pipe = [
        {"$match": {"company_id": cid, "status": {"$ne": "DEPLETED"}}},
        {"$group": {"_id": None, "v": {"$sum": {"$multiply": ["$quantity_left", "$unit_price"]}}}},
    ]
    legacy_val = await db.material_lots.aggregate(legacy_val_pipe).to_list(1)
    stock_value = round(
        (loc_val[0]["v"] if loc_val else 0) + (legacy_val[0]["v"] if legacy_val else 0), 2
    )

    stock = await warehouse_stock(user)  # type: ignore
    low = [s for s in stock if s.get("low_stock")]

    # Top movements this week — by material (works for both location and lot movements)
    week_start = now_utc() - timedelta(days=7)
    top_pipe = [
        {"$match": {"company_id": cid, "timestamp": {"$gte": week_start}}},
        {"$group": {"_id": "$material_id", "n": {"$sum": 1}, "last_code": {"$last": "$material_code"}}},
        {"$sort": {"n": -1}},
        {"$limit": 10},
    ]
    tops = await db.lot_movements.aggregate(top_pipe).to_list(10)
    top_materials = []
    for t in tops:
        mid = t.get("_id")
        if not mid:
            continue
        m = await db.materials.find_one({"id": mid}, {"_id": 0, "name": 1, "code": 1})
        top_materials.append({
            "lot_code": t.get("last_code") or (m or {}).get("code") or "—",
            "material_name": (m or {}).get("name"),
            "movements": t["n"],
        })

    return {
        "lots_count": lots_count,
        "movements_today": movements_today,
        "low_stock_count": len(low),
        "stock_value": stock_value if user["role"] == "ADMIN" else None,
        "low_stock": low[:10],
        "top_movements": top_materials,
    }


def _build_label_png(lot: dict, material: dict, zone_name: Optional[str] = None) -> bytes:
    """Build a 80mm-wide PNG label using PIL + qrcode."""
    import qrcode
    from PIL import Image, ImageDraw, ImageFont
    from io import BytesIO
    width = 600  # ~80mm at 192dpi-ish
    height = 760
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    try:
        font_b = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 36)
        font_m = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 26)
        font_r = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 22)
        font_s = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 18)
    except Exception:
        font_b = font_m = font_r = font_s = ImageFont.load_default()
    # Header
    draw.text((width // 2, 30), "ELEGANT GLASS", anchor="mm", fill="black", font=font_b)
    draw.line([(20, 70), (width - 20, 70)], fill="black", width=2)
    # QR
    qr = qrcode.QRCode(box_size=8, border=1)
    qr.add_data(lot["lot_code"]); qr.make(fit=True)
    qimg = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    qw, qh = qimg.size
    img.paste(qimg, ((width - qw) // 2, 90))
    # Lot code below QR
    draw.text((width // 2, 90 + qh + 18), lot["lot_code"], anchor="mm", fill="black", font=font_m)
    y = 90 + qh + 60
    # Material name (wrap)
    name = (material or {}).get("name", "—")
    if len(name) > 28:
        draw.text((30, y), name[:28], fill="black", font=font_b); y += 42
        draw.text((30, y), name[28:56], fill="black", font=font_b); y += 42
    else:
        draw.text((30, y), name, fill="black", font=font_b); y += 44
    # Quantity + entry date
    qty_line = f"{lot['quantity']:g} {(material or {}).get('unit', '')} · Entrada {lot['entry_date'].strftime('%d/%m/%Y') if isinstance(lot['entry_date'], datetime) else str(lot['entry_date'])[:10]}"
    draw.text((30, y), qty_line, fill="black", font=font_r); y += 32
    if lot.get("supplier_name"):
        draw.text((30, y), f"Proveedor: {lot['supplier_name']}", fill="black", font=font_s); y += 28
    if zone_name:
        draw.text((30, y), f"Zona: {zone_name} · {lot.get('row_label') or ''}".strip(), fill="black", font=font_s); y += 28
    # Footer
    draw.line([(20, height - 50), (width - 20, height - 50)], fill="black", width=2)
    draw.text((width // 2, height - 25), "glasswork.app", anchor="mm", fill="black", font=font_s)
    buf = BytesIO(); img.save(buf, format="PNG")
    return buf.getvalue()


@api.get("/warehouse/lots/{lot_code}/label.png")
async def warehouse_label_png(lot_code: str, user: dict = Depends(get_current_user)):
    from fastapi.responses import Response
    lot = await db.material_lots.find_one({"lot_code": lot_code, "company_id": user["company_id"]})
    if not lot: raise HTTPException(status_code=404, detail="Lote no encontrado")
    mat = await db.materials.find_one({"id": lot["material_id"]}, {"_id": 0})
    zone_name = None
    if lot.get("zone_id"):
        z = await db.storage_zones.find_one({"id": lot["zone_id"]}, {"_id": 0, "name": 1})
        zone_name = (z or {}).get("name")
    png = _build_label_png(lot, mat or {}, zone_name)
    return Response(content=png, media_type="image/png")


@api.get("/warehouse/lots/{lot_code}/label-preview")
async def warehouse_label_preview(lot_code: str, user: dict = Depends(get_current_user)):
    """Returns base64 PNG for in-app preview."""
    lot = await db.material_lots.find_one({"lot_code": lot_code, "company_id": user["company_id"]})
    if not lot: raise HTTPException(status_code=404, detail="Lote no encontrado")
    mat = await db.materials.find_one({"id": lot["material_id"]}, {"_id": 0})
    zone_name = None
    if lot.get("zone_id"):
        z = await db.storage_zones.find_one({"id": lot["zone_id"]}, {"_id": 0, "name": 1})
        zone_name = (z or {}).get("name")
    png = _build_label_png(lot, mat or {}, zone_name)
    import base64 as _b64
    return {
        "filename": f"etiqueta_{lot_code}.png",
        "mime": "image/png",
        "base64": _b64.b64encode(png).decode("ascii"),
    }


def _build_escpos_bytes(lot: dict, material: dict, zone_name: Optional[str]) -> bytes:
    """Build ESC/POS sequence for an 80mm thermal printer."""
    import qrcode
    ESC = b"\x1b"; GS = b"\x1d"
    out = bytearray()
    out += ESC + b"@"  # init
    out += ESC + b"a" + b"\x01"  # center
    # Title double height + bold
    out += ESC + b"E" + b"\x01"  # bold on
    out += GS + b"!" + b"\x11"  # double width+height
    out += b"ELEGANT GLASS\n"
    out += GS + b"!" + b"\x00"  # normal
    out += ESC + b"E" + b"\x00"  # bold off
    out += b"--------------------------------\n"
    # QR using GS ( k commands (model 2)
    code = lot["lot_code"].encode("ascii")
    # Set model 2
    out += GS + b"(k" + b"\x04\x00\x31\x41\x32\x00"
    # Set size (1-16); 8 = large
    out += GS + b"(k" + b"\x03\x00\x31\x43" + bytes([8])
    # Error correction level L
    out += GS + b"(k" + b"\x03\x00\x31\x45\x30"
    # Store data
    pl = len(code) + 3
    out += GS + b"(k" + bytes([pl & 0xff, (pl >> 8) & 0xff]) + b"\x31\x50\x30" + code
    # Print
    out += GS + b"(k" + b"\x03\x00\x31\x51\x30"
    out += b"\n"
    # Lot code text (bold)
    out += ESC + b"E" + b"\x01"
    out += GS + b"!" + b"\x10"  # double width
    out += code + b"\n"
    out += GS + b"!" + b"\x00"
    out += ESC + b"E" + b"\x00"
    out += ESC + b"a" + b"\x00"  # left
    out += b"\n"
    name = (material or {}).get("name", "")[:60]
    out += ESC + b"E" + b"\x01" + name.encode("utf-8", "replace") + b"\n" + ESC + b"E" + b"\x00"
    qty_line = f"{lot['quantity']:g} {(material or {}).get('unit','')}"
    entry = lot["entry_date"]
    if isinstance(entry, datetime):
        date_s = entry.strftime("%d/%m/%Y")
    else:
        date_s = str(entry)[:10]
    out += f"{qty_line} - Entrada {date_s}\n".encode("utf-8", "replace")
    if lot.get("supplier_name"):
        out += f"Proveedor: {lot['supplier_name']}\n".encode("utf-8", "replace")
    if zone_name:
        out += f"Zona: {zone_name} {lot.get('row_label') or ''}\n".encode("utf-8", "replace")
    out += b"--------------------------------\n"
    out += b"\n\n\n"
    out += GS + b"V" + b"\x42" + b"\x00"  # partial cut
    return bytes(out)


@api.post("/warehouse/lots/{lot_code}/print")
async def warehouse_print(lot_code: str, user: dict = Depends(get_current_user)):
    lot = await db.material_lots.find_one({"lot_code": lot_code, "company_id": user["company_id"]})
    if not lot: raise HTTPException(status_code=404, detail="Lote no encontrado")
    mat = await db.materials.find_one({"id": lot["material_id"]}, {"_id": 0})
    zone_name = None
    if lot.get("zone_id"):
        z = await db.storage_zones.find_one({"id": lot["zone_id"]}, {"_id": 0, "name": 1})
        zone_name = (z or {}).get("name")
    payload = _build_escpos_bytes(lot, mat or {}, zone_name)
    printer_ip = os.environ.get("PRINTER_IP", "").strip()
    printer_port = int(os.environ.get("PRINTER_PORT", "9100") or 9100)
    if not printer_ip:
        raise HTTPException(status_code=503, detail="Impresora no configurada (PRINTER_IP vacío). Configura la IP en .env y reintenta.")
    import socket as _socket
    try:
        with _socket.create_connection((printer_ip, printer_port), timeout=5) as s:
            s.sendall(payload)
        return {"ok": True, "bytes": len(payload)}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"No se pudo conectar a la impresora ({printer_ip}:{printer_port}): {e}")


async def _try_send_escpos(payload: bytes) -> dict:
    """Try to send ESC/POS payload to the thermal printer. Returns {printed, message}.
    Never raises — if printer is not reachable we report printed=false gracefully."""
    printer_ip = os.environ.get("PRINTER_IP", "").strip()
    printer_port = int(os.environ.get("PRINTER_PORT", "9100") or 9100)
    if not printer_ip:
        return {
            "printed": False,
            "printer_configured": False,
            "message": "Impresora no configurada aún. La etiqueta se ha generado y está lista para imprimir cuando conectes la impresora (PRINTER_IP).",
            "bytes": len(payload),
        }
    import socket as _socket
    try:
        with _socket.create_connection((printer_ip, printer_port), timeout=4) as s:
            s.sendall(payload)
        return {
            "printed": True,
            "printer_configured": True,
            "message": f"Etiqueta enviada a la impresora {printer_ip}.",
            "bytes": len(payload),
        }
    except Exception as e:
        return {
            "printed": False,
            "printer_configured": True,
            "message": f"No se pudo conectar a la impresora ({printer_ip}:{printer_port}): {e}",
            "bytes": len(payload),
        }


class AssignAndPrintIn(BaseModel):
    lot_code: str


@api.post("/warehouse/assign-and-print")
async def warehouse_assign_and_print(
    body: AssignAndPrintIn,
    request: Request,
    user: dict = Depends(get_current_user),
):
    """Automatic classification + print flow.

    1. Identify the lot and its material category.
    2. Find the first storage zone whose category matches the material's.
    3. Assign the first row with available capacity (<= 6 lots per row).
    4. Persist the zone_id + row_label on the lot and record a LOCATE movement.
    5. Build an ESC/POS payload and try to send it to the thermal printer. If
       PRINTER_IP is not configured (or unreachable), we still return 200 with
       printed=false so the UI can show "etiqueta preparada, imprimir luego".
    """
    lot = await db.material_lots.find_one({"lot_code": body.lot_code, "company_id": user["company_id"]})
    if not lot:
        raise HTTPException(status_code=404, detail="Lote no encontrado")
    mat = await db.materials.find_one({"id": lot["material_id"]}, {"_id": 0}) or {}
    category = mat.get("category")
    if not category:
        raise HTTPException(status_code=422, detail="El material del lote no tiene categoría, no se puede clasificar.")

    # 1) Find matching zone — prefer zones with available capacity
    zones = await db.storage_zones.find(
        {"company_id": user["company_id"], "category": category}
    ).sort("name", 1).to_list(200)
    if not zones:
        raise HTTPException(
            status_code=404,
            detail=f"No hay ninguna zona configurada para la categoría {category}. Crea una zona en Almacén > Zonas primero.",
        )

    # Config — max lots per row inside a zone
    MAX_PER_ROW = int(os.environ.get("WAREHOUSE_MAX_PER_ROW", "6") or 6)

    chosen_zone = None
    chosen_row = None
    for z in zones:
        row_count = int(z.get("row_count") or 10)
        # Count lots already located in each row of this zone (excluding depleted)
        lots_in_zone = await db.material_lots.find({
            "company_id": user["company_id"],
            "zone_id": z["id"],
            "status": {"$ne": "DEPLETED"},
            "id": {"$ne": lot["id"]},  # exclude current lot so re-scanning keeps its slot
        }, {"row_label": 1}).to_list(2000)
        row_counts: dict = {}
        for existing in lots_in_zone:
            rl = (existing.get("row_label") or "").strip()
            if rl:
                row_counts[rl] = row_counts.get(rl, 0) + 1
        # Walk rows in order Fila 1 … Fila N and pick the first with capacity
        for i in range(1, row_count + 1):
            rl = f"Fila {i}"
            if row_counts.get(rl, 0) < MAX_PER_ROW:
                chosen_zone = z
                chosen_row = rl
                break
        if chosen_zone:
            break

    if not chosen_zone:
        # All zones for this category are full → fall back to first zone, row 1
        chosen_zone = zones[0]
        chosen_row = "Fila 1"

    # 2) Persist assignment
    previous_zone = lot.get("zone_id")
    await db.material_lots.update_one(
        {"id": lot["id"]},
        {"$set": {
            "zone_id": chosen_zone["id"],
            "row_label": chosen_row,
            "updated_at": now_utc(),
        }},
    )

    # 3) Record LOCATE movement
    mv = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "lot_id": lot["id"],
        "type": "LOCATE",
        "quantity": None,
        "worker_id": user["id"],
        "worker_name": user.get("name"),
        "project_id": None,
        "project_name": None,
        "timestamp": now_utc(),
        "note": (
            f"Clasificación automática → {chosen_zone['name']} · {chosen_row}"
            + (" (reubicado)" if previous_zone and previous_zone != chosen_zone["id"] else "")
        ),
    }
    await db.lot_movements.insert_one(mv.copy())

    # 4) Build ESC/POS and try print
    lot_reloaded = await db.material_lots.find_one({"id": lot["id"]}, {"_id": 0})
    payload = _build_escpos_bytes(lot_reloaded or lot, mat, chosen_zone["name"])
    print_result = await _try_send_escpos(payload)

    await audit_log(
        db, action="WAREHOUSE_AUTO_CLASSIFY", resource="lot", resource_id=lot["id"],
        request=request, user=user,
        extra={
            "lot_code": lot["lot_code"],
            "zone": chosen_zone["name"],
            "row_label": chosen_row,
            "category": category,
            "relocated": bool(previous_zone and previous_zone != chosen_zone["id"]),
            "printed": print_result.get("printed"),
        },
    )

    return {
        "ok": True,
        "lot": {
            "id": lot["id"],
            "lot_code": lot["lot_code"],
            "quantity_left": lot.get("quantity_left"),
        },
        "material": {
            "id": mat.get("id"),
            "name": mat.get("name"),
            "category": category,
            "unit": mat.get("unit"),
        },
        "zone": {
            "id": chosen_zone["id"],
            "name": chosen_zone["name"],
            "category": chosen_zone["category"],
        },
        "row_label": chosen_row,
        "relocated": bool(previous_zone and previous_zone != chosen_zone["id"]),
        "print": print_result,
    }


# =========================================================
# Health
# =========================================================
@api.get("/")
async def root():
    return {"app": "GLASSWORK", "ok": True}


# =========================================================
# Security — admin-only audit log viewer
# =========================================================
@api.get("/security/audit-logs")
async def list_audit_logs(
    action: Optional[str] = None,
    user_email: Optional[str] = None,
    limit: int = Query(100, le=500),
    user: dict = Depends(require_role("ADMIN")),
):
    flt: dict = {"company_id": user["company_id"]}
    if action: flt["action"] = action
    if user_email: flt["user_email"] = {"$regex": user_email, "$options": "i"}
    docs = await db.audit_logs.find(flt, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    for d in docs:
        if isinstance(d.get("timestamp"), datetime):
            d["timestamp"] = iso(d["timestamp"])
    return docs


@api.get("/security/sessions")
async def list_active_sessions(user: dict = Depends(require_role("ADMIN"))):
    """List active access-token sessions for the current company's users."""
    user_ids = [u["id"] async for u in db.users.find({"company_id": user["company_id"]}, {"id": 1})]
    docs = await db.token_sessions.find(
        {"user_id": {"$in": user_ids}}, {"_id": 0}
    ).sort("last_used", -1).to_list(500)
    for d in docs:
        for k in ("issued_at", "last_used", "exp"):
            if isinstance(d.get(k), datetime):
                d[k] = iso(d[k])
    return docs


# =========================================================
# Seeding
# =========================================================
SPANISH_MATERIALS = [
    # (name, unit, category, unit_price, supplier)
    ("Perfil Cortizo Serie 70 Marco", "m", "PERFILERIA", 18.50, "Cortizo"),
    ("Perfil Cortizo Serie 70 Hoja", "m", "PERFILERIA", 19.20, "Cortizo"),
    ("Perfil Cortizo COR-60 Marco", "m", "PERFILERIA", 14.80, "Cortizo"),
    ("Perfil Technal Soleal", "m", "PERFILERIA", 22.10, "Technal"),
    ("Perfil corredera Schüco ASS 50", "m", "PERFILERIA", 25.40, "Schüco"),
    ("Perfil oscilobatiente Reynaers", "m", "PERFILERIA", 27.90, "Reynaers"),
    ("Junquillo aluminio anodizado", "m", "PERFILERIA", 3.20, "Cortizo"),
    ("Tapa Junta Cortizo", "m", "PERFILERIA", 2.80, "Cortizo"),
    ("Vidrio laminar 6+6 incoloro", "m2", "VIDRIO", 68.00, "Saint-Gobain"),
    ("Vidrio laminar 8+8 acústico", "m2", "VIDRIO", 95.00, "Guardian"),
    ("Vidrio templado 10mm", "m2", "VIDRIO", 78.00, "Saint-Gobain"),
    ("Vidrio Climalit 4/16/4 bajo emisivo", "m2", "VIDRIO", 72.50, "Saint-Gobain"),
    ("Vidrio Climalit 6/20/6 control solar", "m2", "VIDRIO", 92.00, "Guardian"),
    ("Vidrio espejo 5mm", "m2", "VIDRIO", 35.00, "Saint-Gobain"),
    ("Vidrio mate ácido 6mm", "m2", "VIDRIO", 56.00, "Saint-Gobain"),
    ("Vidrio templado curvo", "m2", "VIDRIO", 145.00, "Cricursa"),
    ("Herraje corredera Roto Patio", "ud", "HERRAJES", 145.00, "Roto"),
    ("Herraje oscilobatiente Roto NT", "ud", "HERRAJES", 85.00, "Roto"),
    ("Cierre embutido Giesse", "ud", "HERRAJES", 32.50, "Giesse"),
    ("Manilla Hoppe Atlanta", "ud", "HERRAJES", 28.40, "Hoppe"),
    ("Manilla con llave Tesa", "ud", "HERRAJES", 45.00, "Tesa"),
    ("Bisagra oculta Anuba", "ud", "HERRAJES", 18.00, "Anuba"),
    ("Compás limitador apertura", "ud", "HERRAJES", 22.00, "Roto"),
    ("Cremona corredera elevable", "ud", "HERRAJES", 95.00, "GU"),
    ("Cerradura embutir multipunto", "ud", "HERRAJES", 165.00, "Tesa"),
    ("Silicona neutra Bostik N3500", "ud", "SELLANTES", 7.20, "Bostik"),
    ("Silicona estructural Sika SG500", "ud", "SELLANTES", 14.50, "Sika"),
    ("Silicona acética transparente", "ud", "SELLANTES", 5.80, "Pattex"),
    ("Espuma de poliuretano profesional", "ud", "SELLANTES", 9.50, "Sika"),
    ("Cinta butilo doble cara", "m", "SELLANTES", 4.20, "3M"),
    ("Junta EPDM marco", "m", "SELLANTES", 1.80, "Hutchinson"),
    ("Junta acristalamiento perimetral", "m", "SELLANTES", 2.10, "Hutchinson"),
    ("Imprimación silicona", "ud", "SELLANTES", 18.00, "Sika"),
    ("Taladro percutor Bosch GSB 18V", "ud", "HERRAMIENTAS", 220.00, "Bosch"),
    ("Pistola silicona profesional", "ud", "HERRAMIENTAS", 38.00, "Bostik"),
    ("Ventosa doble vidrio 80kg", "ud", "HERRAMIENTAS", 145.00, "Bohle"),
    ("Cortador vidrio ruleta", "ud", "HERRAMIENTAS", 22.00, "Bohle"),
    ("Nivel láser autonivelante", "ud", "HERRAMIENTAS", 285.00, "Bosch"),
    ("Calzo regulable PVC", "caja", "CONSUMIBLES", 12.00, "Reca"),
    ("Tornillo autoperforante 4.8x25", "caja", "CONSUMIBLES", 8.50, "Reca"),
    ("Tornillo cabeza avellanada 4x40", "caja", "CONSUMIBLES", 6.20, "Reca"),
    ("Taco metálico expansión", "caja", "CONSUMIBLES", 14.00, "Fischer"),
    ("Taco nylon SX 8", "caja", "CONSUMIBLES", 9.40, "Fischer"),
    ("Disco corte aluminio", "ud", "CONSUMIBLES", 4.80, "Bosch"),
    ("Disco corte vidrio diamante", "ud", "CONSUMIBLES", 38.00, "Bosch"),
    ("Bayeta microfibra", "caja", "CONSUMIBLES", 18.00, "Vileda"),
    ("Limpiador vidrio profesional", "l", "CONSUMIBLES", 6.50, "Bohle"),
    ("Cinta señalización obra", "ud", "OTROS", 4.20, "Reca"),
    ("Casco protección EPI", "ud", "OTROS", 22.00, "3M"),
    ("Guantes anticorte vidrio", "ud", "OTROS", 8.50, "Mapa"),
]


SAMPLE_PHOTO_BASE64 = (
    "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR4nGNgYGD4DwABBAEAfbLI3wAAAABJRU5ErkJggg=="
)


async def ensure_indexes():
    await db.users.create_index("email", unique=True)
    await db.users.create_index("company_id")
    await db.projects.create_index("company_id")
    await db.material_entries.create_index([("company_id", 1), ("project_id", 1)])
    await db.daily_logs.create_index([("company_id", 1), ("worker_id", 1)])
    await db.project_photos.create_index([("company_id", 1), ("project_id", 1)])
    await db.material_lots.create_index("lot_code", unique=True)
    await db.material_lots.create_index([("company_id", 1), ("zone_id", 1)])
    await db.lot_movements.create_index([("company_id", 1), ("lot_id", 1)])
    await db.storage_zones.create_index([("company_id", 1), ("name", 1)])






@app.on_event("startup")
async def startup():
    await ensure_indexes()
    # Security layer setup
    await ensure_security_indices(db)
    purged = await cleanup_expired_security_records(db)
    logger.info(f"Security cleanup on startup: {purged}")
    # Push notifications indices
    await ensure_notification_indices(db)


@app.on_event("shutdown")
async def shutdown():
    client.close()


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
# Security headers — added LAST so it runs first on response (Starlette stacks LIFO)
app.add_middleware(SecurityHeadersMiddleware)
